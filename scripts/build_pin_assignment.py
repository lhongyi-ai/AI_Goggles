#!/usr/bin/env python3
"""Build all pin-assignment artifacts from the working master (cm4_pinmap.py).

Working format -> delivery format:
  scripts/cm4_pinmap.py  (master)
      ├─> config/cm4_v1_pin_assignment.yaml            (gate-readable, working)
      ├─> generated/reports/cm4_v1_pin_assignment.csv  (review/diff, working)
      └─> generated/reports/cm4_v1_pin_assignment.xlsx (formal delivery)

Run after editing cm4_pinmap.py, then run scripts/check_pin_freeze.py.
"""
from __future__ import annotations

import csv
from pathlib import Path

import yaml

try:
    from .cm4_pinmap import ASSIGNMENTS, METADATA, OPEN_TBD
    from .kicad_tools import PROJECT_ROOT, status
except ImportError:
    from cm4_pinmap import ASSIGNMENTS, METADATA, OPEN_TBD
    from kicad_tools import PROJECT_ROOT, status

YAML_FILE = PROJECT_ROOT / "config" / "cm4_v1_pin_assignment.yaml"
CSV_FILE = PROJECT_ROOT / "generated" / "reports" / "cm4_v1_pin_assignment.csv"
XLSX_FILE = PROJECT_ROOT / "generated" / "reports" / "cm4_v1_pin_assignment.xlsx"

COLUMNS = [
    "function", "signal", "connector", "pin", "cm4_net", "cm4_func",
    "voltage_domain", "pinmux_conflict", "board_connects_to",
    "device_tree_status", "priority", "source_verified",
]


def write_yaml() -> None:
    # Gate checker reads: status, assignments[].(required fields), open_tbd[].
    doc = {
        "status": METADATA["status"],
        "frozen_date": None,
        "frozen_by": None,
        "schematic_ref": METADATA["schematic_ref"],
        "pinout_ref": METADATA["pinout_ref"],
        "soc_ref": METADATA["soc_ref"],
        "gpio_vref": METADATA["gpio_vref"],
        "assignments": ASSIGNMENTS,
        "open_tbd": OPEN_TBD,
        "gate": {
            "may_enter_layout": False,
            "reason": "Pins verified vs CM4 V1.20; awaiting procurement TBD freeze",
        },
    }
    header = (
        "# ═══════════════════════════════════════════════════════════════════\n"
        "# CM4 V1 PIN ASSIGNMENT — FREEZE GATE  (GENERATED FILE — DO NOT EDIT)\n"
        "# Source of truth: scripts/cm4_pinmap.py  |  Rebuild: build_pin_assignment.py\n"
        f"# Pins verified against {METADATA['pinout_ref']} + {METADATA['schematic_ref']}.\n"
        "# ═══════════════════════════════════════════════════════════════════\n"
    )
    with YAML_FILE.open("w", encoding="utf-8") as fh:
        fh.write(header)
        yaml.safe_dump(doc, fh, sort_keys=False, allow_unicode=True, width=100)


def write_csv() -> None:
    CSV_FILE.parent.mkdir(parents=True, exist_ok=True)
    with CSV_FILE.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in ASSIGNMENTS:
            writer.writerow(row)


def write_xlsx() -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Sheet 1: cover / metadata
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "AI Glasses Carrier V1 — CM4 Pin Assignment"
    cover["A1"].font = Font(size=14, bold=True)
    meta_rows = [
        ("Board", METADATA["board"]),
        ("Status", METADATA["status"]),
        ("Schematic ref", METADATA["schematic_ref"]),
        ("Pinout ref", METADATA["pinout_ref"]),
        ("SoC ref", METADATA["soc_ref"]),
        ("GPIO_VREF", METADATA["gpio_vref"]),
        ("Total rows", len(ASSIGNMENTS)),
        ("P0 rows", sum(1 for r in ASSIGNMENTS if r["priority"] == "P0")),
        ("Note", "Pins verified vs official pinout. Procurement TBDs still gate full FROZEN."),
    ]
    for i, (k, v) in enumerate(meta_rows, start=3):
        cover[f"A{i}"] = k
        cover[f"A{i}"].font = Font(bold=True)
        cover[f"B{i}"] = str(v)
    cover.column_dimensions["A"].width = 16
    cover.column_dimensions["B"].width = 70

    # Sheet 2: assignments
    ws = wb.create_sheet("Pin Assignment")
    headers = [c.replace("_", " ").title() for c in COLUMNS]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    p0_fill = PatternFill("solid", fgColor="E2EFDA")
    p1_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in ASSIGNMENTS:
        ws.append([row.get(c, "") for c in COLUMNS])
        fill = p0_fill if row["priority"] == "P0" else p1_fill
        for col in range(1, len(COLUMNS) + 1):
            ws.cell(row=ws.max_row, column=col).fill = fill

    widths = [11, 26, 11, 9, 26, 24, 20, 26, 30, 26, 9, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    # Sheet 3: open TBD
    tbd = wb.create_sheet("Open TBD (8.3)")
    tbd.append(["ID", "Must confirm", "Resolved"])
    for col in range(1, 4):
        c = tbd.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
    for item in OPEN_TBD:
        tbd.append([item["id"], item["must_confirm"], "YES" if item["resolved"] else "NO"])
    tbd.column_dimensions["A"].width = 16
    tbd.column_dimensions["B"].width = 70
    tbd.column_dimensions["C"].width = 10

    XLSX_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_FILE)


def main() -> int:
    write_yaml()
    write_csv()
    write_xlsx()
    print(status("OK", f"Working YAML (gate): {YAML_FILE}"))
    print(status("OK", f"Working CSV:         {CSV_FILE}"))
    print(status("OK", f"DELIVERY XLSX:       {XLSX_FILE}"))
    print(status("INFO", f"{len(ASSIGNMENTS)} rows written; status={METADATA['status']}."))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
