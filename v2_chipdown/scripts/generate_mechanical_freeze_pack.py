#!/usr/bin/env python3
"""Generate the V2.3 mechanical-freeze audit pack.

Phase 1 is audit-only: it does not edit the KiCad schematic, footprints, or PCB.
It turns the current schematic/BOM plus the available local supplier documents
into one mechanical database and a reproducible set of review artifacts.
"""
from __future__ import annotations

import csv
import math
import os
import sys
import textwrap
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

os.environ.setdefault("MPLCONFIGDIR", "/private/tmp/ai_glasses_v2_mplcache")

import matplotlib

matplotlib.use("Agg")
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
from matplotlib import patches
from matplotlib.backends.backend_pdf import PdfPages

ROOT = Path(__file__).resolve().parents[2]
V2 = ROOT / "v2_chipdown"
CFG = V2 / "config"
REPORTS = V2 / "reports"
OUT = REPORTS / "output"
FIGS = OUT / "mechanical_freeze"
PHASE = REPORTS / "phase_reports"
REVIEWS = REPORTS / "engineering_reviews"

sys.path.insert(0, str(V2 / "scripts"))
from chipdown_bom import COMPONENTS, _validate, all_nets, net_endpoints, net_meta  # noqa: E402


TODAY = date.today().isoformat()
DIM_PDF_V22 = "/Users/stanley/Downloads/AI_Glasses_Hardware_Dimensions_V2.2 (3).pdf"
DIM_PDF_V23 = "/Users/stanley/Downloads/AI_Glasses_Hardware_Dimensions_V2.3.pdf"


FONT_FAMILY = "Arial Unicode MS"
try:
    fm.findfont(FONT_FAMILY, fallback_to_default=False)
except Exception:
    FONT_FAMILY = "PingFang SC"

plt.rcParams.update(
    {
        "font.family": FONT_FAMILY,
        "axes.unicode_minus": False,
        "pdf.fonttype": 42,
        "ps.fonttype": 42,
    }
)


STATUS_COLORS = {
    "FROZEN": "92D050",
    "PROVISIONAL": "FFE699",
    "RANGE": "F4B183",
    "HOLD": "F8CBAD",
    "TBD": "D9E1F2",
    "DNP": "D9D9D9",
}


@dataclass(frozen=True)
class Source:
    doc: str
    title: str
    revision: str
    page: str
    url_or_file: str
    confidence: str


SOURCES: dict[str, Source] = {
    "schematic": Source(
        "S0",
        "Current V2 generated schematic/BOM master",
        f"generated {TODAY}",
        "scripts/chipdown_bom.py",
        "v2_chipdown/scripts/chipdown_bom.py",
        "high for current project intent; not a package drawing",
    ),
    "dim_v22": Source(
        "S1",
        "AI Glasses Hardware Dimensions V2.2",
        "2026-07-01",
        "pp. 1-10",
        DIM_PDF_V22,
        "medium; project extraction from BOM PDFs",
    ),
    "dim_v23": Source(
        "S2",
        "AI Glasses Hardware Dimensions V2.3",
        "2026-07-01",
        "pp. 1-10",
        DIM_PDF_V23,
        "medium; project extraction from BOM PDFs",
    ),
    "fcu760k": Source(
        "S3",
        "Quectel FCU760K Short Specification",
        "V1.1, Released, 2024",
        "pp. 1-3",
        "AI_Glasses_HOLD_Closure_Pack/02_WiFi_FCU760K/Quectel_FCU760K_Short_Spec_V1.1.pdf",
        "high for module outline and basic electrical data; land pattern still missing",
    ),
    "ndp120": Source(
        "S4",
        "Syntiant NDP120 Product Brief",
        "V1, 2023",
        "pp. 1-2",
        "AI_Glasses_HOLD_Closure_Pack/01_NDP120/NDP120_Product_Brief.pdf",
        "medium; product brief, not full datasheet/package drawing",
    ),
    "imx415": Source(
        "S5",
        "Sony IMX415-AAQR-C Datasheet E19504",
        "Official edition, 2019-05-21",
        "pp. 1, 3, 97",
        "AI_Glasses_HOLD_Closure_Pack/04_IMX415/IMX415-AAQR-C_Datasheet_E19504.pdf",
        "high for bare sensor; low for final module mechanics",
    ),
    "bq2970": Source(
        "S6",
        "TI BQ297xx Datasheet SLUSBU9I",
        "March 2014, revised August 2024",
        "pp. 1, 3, package outline",
        "AI_Glasses_HOLD_Closure_Pack/03_Battery_LP451165_1S2P/TI_BQ2970_1S_Protection_Datasheet.pdf",
        "high for IC; exact suffix still TBD",
    ),
    "battery_fit": Source(
        "S7",
        "LP451165 Battery Bay Fit Check",
        "project fit-check, 2026-07-01",
        "dimensions.csv + fit-check md",
        "AI_Glasses_HOLD_Closure_Pack/06_Mechanical_Fit_Check/",
        "medium; CAD clearance envelope, not supplier pack drawing",
    ),
    "rk3576": Source(
        "S8",
        "RK3576 Brief Datasheet file",
        "V1.2, 2024-03-11; PDF metadata title mismatch noted",
        "pp. 1-3",
        "AI_Glasses_HOLD_Closure_Pack/05_Project_References/RK3576 Brief Datasheet V1.2-20240311.pdf",
        "medium; package/HDG/ball map still required",
    ),
}


def source_dict(key: str) -> dict[str, str]:
    s = SOURCES[key]
    return {
        "source_document": s.title,
        "source_revision": s.revision,
        "source_page": s.page,
        "source_url_or_file": s.url_or_file,
        "confidence": s.confidence,
    }


def d(l: float | None = None, w: float | None = None, h: float | None = None) -> dict[str, float | None]:
    return {"length_mm": l, "width_mm": w, "height_mm": h}


# Mechanical overrides are intentionally conservative. If an exact MPN/package
# drawing is not in hand, the record stays HOLD/TBD even if the schematic says
# "Fit" for electrical architecture.
OVERRIDES: dict[str, dict[str, Any]] = {
    "U1": dict(
        functional_block="main compute SoC",
        manufacturer="Rockchip",
        exact_mpn="RK3576",
        package="FCCSP698L, 0.6 mm pitch",
        body=d(16.1, 17.2, None),
        operating_voltage="RK806S rails from SOC_5V",
        typical_power="included in compute model",
        peak_power="part of AI-burst compute row",
        freeze_status="HOLD",
        source="rk3576",
        notes="Body outline known from project BOM/PDF; full datasheet, HDG, ball map, max height, DDR guide and official land pattern still required.",
    ),
    "U2": dict(
        functional_block="SoC PMIC",
        manufacturer="Rockchip",
        exact_mpn="RK806S-5",
        package="TBD QFN/BGA variant",
        body=d(),
        operating_voltage="SOC_5V input; generates RK3576/DDR rails",
        typical_power="regulation loss included in compute model",
        peak_power="TBD after rail-current budget",
        freeze_status="HOLD",
        source="dim_v23",
        notes="Exact package, height, land pattern and rail sequence are not frozen.",
    ),
    "U3": dict(
        functional_block="system memory",
        manufacturer="TBD",
        exact_mpn="LPDDR4X 4 GB MPN TBD",
        package="TBD BGA",
        body=d(),
        operating_voltage="RK806 DDR rails",
        typical_power="included in compute model",
        peak_power="TBD by selected MPN",
        freeze_status="HOLD",
        source="dim_v23",
        notes="Needs exact AVL MPN, package drawing, ball map and DDR length/escape review.",
    ),
    "U4": dict(
        functional_block="boot/system storage",
        manufacturer="TBD",
        exact_mpn="32 GB eMMC 5.1 MPN TBD",
        package="BGA153 candidate",
        body=d(),
        operating_voltage="VCC/VCCQ from RK806 rails",
        typical_power="included in compute model",
        peak_power="TBD by selected MPN",
        freeze_status="HOLD",
        source="dim_v23",
        notes="Common eMMC sizes are not freeze data; final MPN and package drawing required.",
    ),
    "Y1": dict(
        functional_block="main clock",
        manufacturer="TBD",
        exact_mpn="24 MHz 10 ppm crystal TBD",
        package="3225 or smaller",
        body=d(3.2, 2.5, None),
        operating_voltage="oscillator pins on RK3576",
        typical_power="passive",
        peak_power="passive",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Reference load capacitance/ESR/placement must follow RK3576 HDG.",
    ),
    "U5": dict(
        functional_block="optional boot flash",
        manufacturer="Macronix",
        exact_mpn="MX25U6432F",
        package="USON/XSON/SOP option",
        body=d(),
        operating_voltage="1.8 V SPI/QSPI",
        typical_power="DNP",
        peak_power="DNP",
        freeze_status="DNP",
        source="dim_v23",
        notes="Footprint kept only if boot flow needs external NOR.",
    ),
    "U6": dict(
        functional_block="main SoC boost",
        manufacturer="Texas Instruments",
        exact_mpn="TPS61088",
        package="VQFN-22 4.5 x 3.5 mm",
        body=d(4.5, 3.5, None),
        operating_voltage="VSYS in, SOC_5V out",
        typical_power="conversion loss in compute-island loss row",
        peak_power="TBD from RK3576 boot/AI current",
        freeze_status="HOLD",
        source="dim_v23",
        notes="Chip size alone is insufficient; L1, input/output caps, current loop and thermal rise are the real footprint.",
    ),
    "L1": dict(
        functional_block="main boost inductor",
        manufacturer="TBD",
        exact_mpn="shielded low-DCR power inductor TBD",
        package="target <=5 x 5 x 2 mm",
        body=d(5.0, 5.0, 2.0),
        operating_voltage="SOC_5V boost switch node",
        typical_power="loss TBD by DCR/RMS current",
        peak_power="Isat must cover boot/AI peak",
        freeze_status="TBD",
        source="dim_v23",
        notes="Target envelope only. Final MPN may grow after Isat/DCR/temperature validation.",
    ),
    "U11": dict(
        functional_block="Wi-Fi/Bluetooth module",
        manufacturer="Quectel",
        exact_mpn="FCU760KAAMD",
        package="LCC module",
        body=d(13.0, 12.2, 2.0),
        mass_g=0.62,
        operating_voltage="VBAT 3.0-3.6 V, typ 3.3 V",
        typical_power="radio average: 80-150 mW in current model",
        peak_power="353 mA @ 3.3 V max Tx current in short spec",
        freeze_status="HOLD",
        source="fcu760k",
        notes="Module outline is known. Official hardware design guide, land pattern/3D and RK3576 driver enumeration still close G05.",
    ),
    "J7": dict(
        functional_block="Wi-Fi antenna",
        manufacturer="TBD",
        exact_mpn="2.4/5 GHz FPC or LDS antenna TBD",
        package="custom antenna/keep-out",
        body=d(),
        operating_voltage="RF passive, 50 ohm",
        typical_power="passive; radiated from U11",
        peak_power="TBD by RF path",
        freeze_status="TBD",
        source="schematic",
        notes="Requires antenna SKU, keep-out, matching and worn-state tuning. Must clear battery, copper, screws, speaker magnet and boost loop.",
    ),
    "U12": dict(
        functional_block="Wi-Fi buck regulator",
        manufacturer="Texas Instruments",
        exact_mpn="TPS62825",
        package="QFN 1.5 x 1.5 mm",
        body=d(1.5, 1.5, None),
        operating_voltage="VSYS to WIFI_3V3",
        typical_power="~5 mW loss allocation",
        peak_power="sized for FCU760K Tx current",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Needs final inductor/cap derating and land-pattern check.",
    ),
    "L2": dict(
        functional_block="Wi-Fi buck inductor",
        manufacturer="TBD",
        exact_mpn="470 nH class inductor TBD",
        package="target 2 x 2 mm class",
        body=d(2.0, 2.0, None),
        operating_voltage="WIFI_3V3 buck",
        typical_power="loss TBD",
        peak_power="sized for FCU760K Tx current",
        freeze_status="TBD",
        source="dim_v23",
        notes="MPN, DCR, Isat and height required.",
    ),
    "U20": dict(
        functional_block="digital class-D speaker amplifier",
        manufacturer="Analog Devices",
        exact_mpn="MAX98360A",
        package="WLP or FC2QFN version TBD",
        body=d(),
        operating_voltage="AUDIO_PWR / I2S control",
        typical_power="audio average 2-3 mW electronics plus speaker output",
        peak_power="depends on speaker impedance and acoustic target",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Exact package and EMI/thermal/acoustic layout still need closure.",
    ),
    "LS1": dict(
        functional_block="main speaker",
        manufacturer="TBD",
        exact_mpn="8 ohm 0.5-1 W micro speaker TBD",
        package="speaker/cavity TBD",
        body=d(),
        operating_voltage="driven by MAX98360A",
        typical_power="13-27 mW average in model",
        peak_power="0.5-1 W acoustic part class",
        freeze_status="TBD",
        source="schematic",
        notes="Must choose speaker MPN, magnet size, port, cavity, leak path and antenna exclusion.",
    ),
    "BT1": dict(
        functional_block="right battery cell",
        manufacturer="cell supplier TBD",
        exact_mpn="LP451165 300 mAh",
        package="pouch cell with tabs/NTC/pack options",
        body=d(70.0, 12.8, 5.6),
        mass_g=6.0,
        operating_voltage="1S Li-Po nominal 3.7 V",
        typical_power="source",
        peak_power="branch current TBD by 1S2P validation",
        freeze_status="HOLD",
        source="battery_fit",
        notes="Use 70 x 12.8 x 5.6 mm CAD clearance envelope, not nominal 65 x 11 x 4.5 mm. Supplier drawing still required.",
    ),
    "BT2": dict(
        functional_block="left battery cell",
        manufacturer="cell supplier TBD",
        exact_mpn="LP451165 300 mAh",
        package="pouch cell with tabs/NTC/pack options",
        body=d(70.0, 12.8, 5.6),
        mass_g=6.0,
        operating_voltage="1S Li-Po nominal 3.7 V",
        typical_power="source",
        peak_power="branch current TBD by 1S2P validation",
        freeze_status="HOLD",
        source="battery_fit",
        notes="Must be lot/capacity/DCIR matched to BT1; one cell per temple, 1S2P.",
    ),
    "J1": dict(
        functional_block="magnetic pogo charge/service",
        manufacturer="TBD",
        exact_mpn="magnetic pogo 4-6 pin TBD",
        package="target about 10 x 3 mm",
        body=d(10.0, 3.0, None),
        operating_voltage="USB_5V, GND, optional USB2/ID",
        typical_power="passive; contact I2R loss",
        peak_power="charge/fault current rating TBD",
        freeze_status="TBD",
        source="dim_v23",
        notes="Needs pin count, contact resistance, magnet polarity, corrosion, short and mating-life data.",
    ),
    "J3": dict(
        functional_block="front FPC connector",
        manufacturer="Hirose",
        exact_mpn="FH26W series pin count TBD",
        package="0.3 mm pitch, 1.0 mm height, bottom-contact/front-flip ZIF",
        body=d(12.0, 3.5, 1.0),
        operating_voltage="MIPI/PDM/I2C/CAM rails/GND",
        typical_power="passive",
        peak_power="pin-current TBD",
        freeze_status="HOLD",
        source="dim_v23",
        notes="35-pin length is only an example. Final pin count and contact orientation are G12 blockers.",
    ),
    "J4": dict(
        functional_block="left-right interconnect / future hinge FPC",
        manufacturer="FPC supplier TBD",
        exact_mpn="custom 6-10 mm FPC / rigid-flex TBD",
        package="0.10-0.18 mm target thickness",
        body={"length_mm": None, "width_mm": "6-10", "height_mm": "0.10-0.18"},
        operating_voltage="BAT_P/VSYS, GND, AON UART, EN/PGOOD/fault/status",
        typical_power="passive; I2R loss",
        peak_power="must carry peak and fault current safely",
        freeze_status="HOLD",
        source="dim_v23",
        notes="EVT-A can be fixed-temple cable/FPC. Folding hinge life, bend radius, stack-up and pin count remain open.",
    ),
    "U7": dict(
        functional_block="always-on BLE MCU",
        manufacturer="Nordic Semiconductor",
        exact_mpn="nRF54L15 QFN48 candidate",
        package="QFN48, 0.4 mm pitch",
        body=d(6.0, 6.0, 0.85),
        operating_voltage="AON_1V8/AON_3V3",
        typical_power="2-8 mW allocation depending state",
        peak_power="radio peak TBD",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="EVT QFN chosen for bring-up. RF layout, SDK and official package drawing still must be checked before fab.",
    ),
    "U8": dict(
        functional_block="AON PMIC/charger/fuel gauge",
        manufacturer="Nordic Semiconductor",
        exact_mpn="nPM1300",
        package="QFN32 5 x 5 mm EVT candidate",
        body=d(5.0, 5.0, None),
        operating_voltage="BAT_P/USB_5V in; AON_1V8/AON_3V3 out",
        typical_power="3-15 mW AON buck/quiescent allocation",
        peak_power="charge/PMIC thermal TBD",
        freeze_status="HOLD",
        source="dim_v23",
        notes="nPM PowerUP config and measured AON <=25/50 mW target must close G08.",
    ),
    "U9": dict(
        functional_block="always-on audio DSP",
        manufacturer="Syntiant",
        exact_mpn="NDP120 QFN40 candidate",
        package="5 x 5 mm 40-pin QFN, 0.4 mm pitch; WLBGA also exists",
        body=d(5.0, 5.0, None),
        operating_voltage="AON rails, reset/wake to nRF54L15",
        typical_power="12-20 mW project listening allocation",
        peak_power="TBD by dev kit measurement",
        freeze_status="HOLD",
        source="ndp120",
        notes="Product brief confirms packages and capability; full datasheet, rail map, reset straps, SDK/licensing and measured listening current still needed.",
    ),
    "U10": dict(
        functional_block="IMU",
        manufacturer="Bosch Sensortec",
        exact_mpn="BMI270",
        package="LGA-14",
        body=d(2.5, 3.0, 0.8),
        operating_voltage="AON_1V8/AON_3V3",
        typical_power="1 mW low-power motion budget",
        peak_power="TBD by IMU mode",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Current schematic uses BMI270. Placement must avoid speaker/motor vibration and high-stress zones.",
    ),
    "U23": dict(
        functional_block="optional 1S battery protection IC",
        manufacturer="Texas Instruments",
        exact_mpn="BQ2970/BQ297xx suffix TBD",
        package="DSE WSON-6",
        body=d(1.5, 1.5, 0.75),
        operating_voltage="single-cell Li-ion protection",
        typical_power="4 uA normal, 100 nA shutdown class",
        peak_power="external FET path handles current",
        freeze_status="DNP",
        source="bq2970",
        notes="DNP if supplier pack has qualified PCM. Exact suffix thresholds must match cell and nPM1300 charger.",
    ),
    "Q1": dict(
        functional_block="optional battery protection MOSFETs",
        manufacturer="TBD",
        exact_mpn="dual back-to-back N-MOSFET TBD",
        package="DFN dual-FET TBD",
        body=d(),
        operating_voltage="cell negative protection path",
        typical_power="DNP",
        peak_power="RDS(on), VDS and pulse current TBD",
        freeze_status="DNP",
        source="schematic",
        notes="Only used with BQ2970 fallback. Must not be populated together with incompatible pack PCM path.",
    ),
    "RS1": dict(
        functional_block="total pack current shunt",
        manufacturer="TBD",
        exact_mpn="10 milliohm 1% 1 W shunt TBD",
        package="1206 shunt candidate",
        body=d(),
        operating_voltage="BAT_P/NPM_VBAT Kelvin sense",
        typical_power="I^2R loss",
        peak_power="sized for pack current",
        freeze_status="PROVISIONAL",
        source="schematic",
        notes="Production-kept sense path; MPN, TCR, surge and pad geometry still need freeze.",
    ),
    "U25": dict(
        functional_block="optional total-current monitor",
        manufacturer="Texas Instruments",
        exact_mpn="INA238",
        package="VSSOP-10",
        body=d(3.0, 4.9, None),
        operating_voltage="AON I2C / shunt monitor",
        typical_power="640 uA class when populated",
        peak_power="5 uA shutdown max class",
        freeze_status="DNP",
        source="dim_v23",
        notes="EVT-A only unless power telemetry stays in production.",
    ),
    "RT1": dict(
        functional_block="right-cell NTC",
        manufacturer="TBD",
        exact_mpn="10k NTC curve TBD",
        package="0402 or film probe",
        body=d(),
        operating_voltage="AON analog / nPM1300 temperature input",
        typical_power="uW class divider",
        peak_power="uW class",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Location matters more than package: place against cell body, not charger IC.",
    ),
    "RT2": dict(
        functional_block="left-cell NTC",
        manufacturer="TBD",
        exact_mpn="10k NTC curve TBD",
        package="0402 or film probe",
        body=d(),
        operating_voltage="AON analog / nPM1300 temperature input",
        typical_power="uW class divider",
        peak_power="uW class",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Needs JEITA/nPM1300 charger config match.",
    ),
    "J6": dict(
        functional_block="BLE antenna",
        manufacturer="TBD",
        exact_mpn="2.4 GHz FPC/PCB antenna TBD",
        package="custom antenna/keep-out",
        body=d(),
        operating_voltage="RF passive, 50 ohm",
        typical_power="passive",
        peak_power="radio-dependent",
        freeze_status="TBD",
        source="schematic",
        notes="Requires left-temple tail keep-out, matching and worn-state tune. Must clear battery/copper/screws/magnets.",
    ),
    "U14": dict(
        functional_block="camera module",
        manufacturer="Sony / module supplier TBD",
        exact_mpn="IMX415-AAQR custom FPC module; Radxa Camera 4K AS001 is only a candidate mechanical baseline",
        package="custom sensor+lens+FPC module",
        body=d(),
        operating_voltage="CAM_1V1, CAM_1V8_SW, CAM_2V9",
        typical_power="250-300 mW camera row in current model",
        peak_power="bare-sensor/module rail currents TBD",
        freeze_status="HOLD",
        source="imx415",
        notes="Current schematic says custom FPC module. The external V2.3 dimensions PDF freezes Radxa AS001 at 32.5 x 32.5 x 18.0 mm, but that is not yet schematic-aligned.",
    ),
    "CAM_LENS": dict(
        functional_block="camera lens stack",
        manufacturer="Radxa/Sony module candidate",
        exact_mpn="Radxa Camera 4K AS001 lens candidate",
        package="M12 x 0.5 lens, approx OD 14 mm",
        body={"length_mm": 32.5, "width_mm": 32.5, "height_mm": 18.0},
        operating_voltage="mechanical/optical",
        typical_power="passive",
        peak_power="passive",
        freeze_status="HOLD",
        source="dim_v23",
        notes="Virtual mechanical item, not in schematic. Must not replace U14 custom module unless EE/CAD approve that architecture.",
    ),
    "MK1": dict(
        functional_block="wake microphone",
        manufacturer="TDK InvenSense",
        exact_mpn="T5837",
        package="bottom-port LGA-5",
        body=d(3.50, 2.65, 0.98),
        operating_voltage="AON microphone rail/PDM",
        typical_power="3 mW wake-mic budget",
        peak_power="TBD by mode",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Requires sound port, acoustic keep-out and wind/AEC placement review.",
    ),
    "MK2": dict(
        functional_block="array microphone 1",
        manufacturer="TDK InvenSense",
        exact_mpn="T5837",
        package="bottom-port LGA-5",
        body=d(3.50, 2.65, 0.98),
        operating_voltage="camera/compute audio PDM rail",
        typical_power="array mic share in active states",
        peak_power="TBD by mode",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Mic coordinate spacing and port path remain G11.",
    ),
    "MK3": dict(
        functional_block="array microphone 2",
        manufacturer="TDK InvenSense",
        exact_mpn="T5837",
        package="bottom-port LGA-5",
        body=d(3.50, 2.65, 0.98),
        operating_voltage="camera/compute audio PDM rail",
        typical_power="array mic share in active states",
        peak_power="TBD by mode",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Mic coordinate spacing and port path remain G11.",
    ),
    "U15": dict(
        functional_block="camera 1.1 V buck",
        manufacturer="Texas Instruments",
        exact_mpn="TPS62840",
        package="WSON/VSON/DSBGA variant TBD",
        body=d(),
        operating_voltage="CAM_1V1",
        typical_power="15-18 mW loss allocation when camera active",
        peak_power="DVDD current plus margin",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Exact package and inductor/cap choices still need freeze.",
    ),
    "L3": dict(
        functional_block="camera buck inductor",
        manufacturer="TBD",
        exact_mpn="cam 1.1 V inductor TBD",
        package="target 1.6 x 0.8 mm class",
        body=d(1.6, 0.8, None),
        operating_voltage="CAM_1V1 buck",
        typical_power="loss TBD",
        peak_power="DVDD current plus ripple",
        freeze_status="TBD",
        source="schematic",
        notes="MPN, DCR, Isat, height and camera noise impact required.",
    ),
    "U16": dict(
        functional_block="camera 2.9 V LDO",
        manufacturer="Texas Instruments",
        exact_mpn="TLV75529PDRVR",
        package="WSON-6 2 x 2 mm",
        body=d(2.0, 2.0, None),
        operating_voltage="CAM_2V9",
        typical_power="14-16 mW loss allocation when camera active",
        peak_power="IMX415 AVDD current plus thermal",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Verify package suffix and PSRR/noise versus final camera module.",
    ),
    "U17": dict(
        functional_block="camera 1.8 V load switch",
        manufacturer="Texas Instruments",
        exact_mpn="TPS22917DBVR",
        package="SOT-23-5",
        body=d(2.9, 2.8, None),
        operating_voltage="AON_1V8 to CAM_1V8_SW",
        typical_power="~1 mW loss allocation",
        peak_power="IOVDD current, reverse-block/QOD check",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Confirm QOD/reverse feed when camera off.",
    ),
    "U18": dict(
        functional_block="MIPI/FPC ESD array 1",
        manufacturer="Texas Instruments",
        exact_mpn="TPD4E05U06",
        package="USON low-cap array",
        body=d(2.5, 1.0, None),
        operating_voltage="MIPI/low-cap ESD",
        typical_power="leakage only",
        peak_power="ESD clamp",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Must be close to FPC entry and not stub MIPI lines.",
    ),
    "U19": dict(
        functional_block="MIPI/FPC ESD array 2",
        manufacturer="Texas Instruments",
        exact_mpn="TPD4E05U06",
        package="USON low-cap array",
        body=d(2.5, 1.0, None),
        operating_voltage="MIPI/low-cap ESD",
        typical_power="leakage only",
        peak_power="ESD clamp",
        freeze_status="PROVISIONAL",
        source="dim_v23",
        notes="Must be close to FPC entry and not stub MIPI lines.",
    ),
    "CAM_FPC": dict(
        functional_block="camera FPC tail",
        manufacturer="camera/FPC supplier TBD",
        exact_mpn="custom camera FPC TBD",
        package="0.2 mm class FPC candidate",
        body=d(),
        operating_voltage="MIPI CSI, I2C, rails, GND",
        typical_power="passive; I2R/insertion loss",
        peak_power="pin-current TBD",
        freeze_status="HOLD",
        source="schematic",
        notes="Virtual mechanical item; final exit direction, stiffener, contact side and bend radius required.",
    ),
}


VIRTUAL_REFS = {"CAM_LENS", "CAM_FPC"}


COMPONENT_CARD_REFS = [
    "U1",
    "U2",
    "U3",
    "U4",
    "U11",
    "U6",
    "L1",
    "U12",
    "U20",
    "Y1",
    "U5",
    "BT1",
    "LS1",
    "J1",
    "J3",
    "J4",
    "U7",
    "U8",
    "U9",
    "U10",
    "U23",
    "Q1",
    "RS1",
    "U25",
    "RT1",
    "BT2",
    "J6",
    "U14",
    "CAM_LENS",
    "MK1",
    "MK2",
    "MK3",
    "U15",
    "U16",
    "U17",
    "U18",
    "CAM_FPC",
]


def comp_by_ref() -> dict[str, Any]:
    out = {c.ref: c for c in COMPONENTS}
    for ref in VIRTUAL_REFS:
        out[ref] = None
    return out


def footprint_readiness(footprint: str, assembly: str) -> str:
    if "VERIFY" in footprint:
        return "NOT_FOR_FAB placeholder; not release-ready"
    if assembly == "DNP":
        return "DNP; do not treat as release BOM"
    if not footprint:
        return "no footprint assigned in KiCad Footprint field; package tracked only in register"
    return "candidate; verify official drawing/land pattern/height before layout"


def record_for(ref: str, comps: dict[str, Any]) -> dict[str, Any]:
    c = comps.get(ref)
    override = OVERRIDES.get(ref, {})
    source_key = override.get("source", "schematic")
    src = source_dict(source_key)
    body = override.get("body", d())
    board = "Virtual mechanical item" if c is None else c.board
    assembly = "HOLD" if c is None else c.assembly
    value = override.get("exact_mpn") if c is None else c.value
    footprint = "" if c is None else c.footprint
    status = override.get("freeze_status") or assembly
    rec = {
        "functional_block": override.get("functional_block", board),
        "reference_designator": ref,
        "bom_id": "VIRTUAL" if c is None else c.bom_id,
        "manufacturer": override.get("manufacturer", "TBD"),
        "exact_mpn": override.get("exact_mpn", value),
        "description": override.get("description", "" if c is None else c.desc),
        "board_region": board,
        "assembly_status": assembly,
        "package": override.get("package", "" if c is None else c.footprint),
        "body_length_mm": body.get("length_mm"),
        "body_width_mm": body.get("width_mm"),
        "body_height_mm": body.get("height_mm"),
        "land_pattern": override.get("land_pattern", footprint),
        "courtyard": override.get("courtyard", "TBD; use official package drawing before layout"),
        "keepout": override.get("keepout", "TBD; see placement notes"),
        "pin_pitch": override.get("pin_pitch", "TBD"),
        "mass_g": override.get("mass_g", "TBD"),
        "operating_voltage": override.get("operating_voltage", "TBD"),
        "typical_power": override.get("typical_power", "TBD"),
        "peak_power": override.get("peak_power", "TBD"),
        "thermal_notes": override.get("thermal_notes", "TBD after thermal review"),
        "placement_notes": override.get("notes", "" if c is None else c.note),
        **src,
        "freeze_status": status,
        "schematic_status": "virtual mechanical note" if c is None else "present in current schematic",
        "footprint_status": "virtual; no KiCad footprint" if c is None else footprint_readiness(footprint, assembly),
        "3d_model_status": override.get("3d_model_status", "HOLD/TBD; no verified aligned STEP in V2 KiCad"),
        "next_action": override.get("next_action", "close gate/source data before PCB layout" if c is None else c.gate),
    }
    return rec


def build_database() -> list[dict[str, Any]]:
    problems = _validate()
    if problems:
        raise RuntimeError("BOM validation failed: " + "; ".join(problems))
    comps = comp_by_ref()
    refs = list(comps.keys())
    records = [record_for(ref, comps) for ref in refs]
    records.sort(key=lambda r: (r["board_region"], r["reference_designator"]))
    return records


def yaml_safe(value: Any) -> Any:
    if isinstance(value, (str, int, float)) or value is None:
        return value
    return str(value)


def write_database(records: list[dict[str, Any]]) -> tuple[Path, Path]:
    CFG.mkdir(parents=True, exist_ok=True)
    yml = CFG / "hardware_mechanical_database.yaml"
    csv_path = CFG / "hardware_mechanical_database.csv"
    payload = {
        "meta": {
            "generated": TODAY,
            "phase": "Phase 1 audit-only mechanical freeze database",
            "source_of_truth": "Current records generated from chipdown_bom.py plus conservative local-source overlays in generate_mechanical_freeze_pack.py",
            "status_legend": ["FROZEN", "PROVISIONAL", "RANGE", "HOLD", "TBD", "DNP"],
            "rule": "No exact MPN/package drawing/land pattern means HOLD or TBD, not release-ready.",
        },
        "sources": {k: source_dict(k) for k in SOURCES},
        "components": [{k: yaml_safe(v) for k, v in r.items()} for r in records],
    }
    yml.write_text(yaml.safe_dump(payload, sort_keys=False, allow_unicode=True), encoding="utf-8")

    fieldnames = list(records[0].keys())
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)
    return yml, csv_path


def md_table(headers: list[str], rows: list[list[Any]]) -> str:
    out = ["| " + " | ".join(headers) + " |", "|" + "|".join(["---"] * len(headers)) + "|"]
    for row in rows:
        out.append("| " + " | ".join(str(x) if x is not None else "TBD" for x in row) + " |")
    return "\n".join(out)


def status_counts(records: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for r in records:
        counts[r["freeze_status"]] = counts.get(r["freeze_status"], 0) + 1
    return counts


def write_design_review(records: list[dict[str, Any]]) -> Path:
    path = REPORTS / "design_review.md"
    rows = [
        [
            "1",
            "左镜腿为什么缺少 side view",
            "HOLD",
            "当前 Phase 1.5 图只有左镜腿 top view，没有左侧厚度 stack 和 A-A/B-B 截面；不能证明 AON 板、电池、FPC、泡棉和壳体在 Z 方向合壳。",
            "补左镜腿 side view、AON tallest-component 截面、电池截面，并绑定数据库尺寸。",
        ],
        [
            "2",
            "右镜腿当前 thickness section 是否完整",
            "PASS WITH CONDITIONS",
            "已有右侧厚度 section，表达了电池不在 RK3576/PMIC/boost 下方；但未含真实最大器件高度、焊锡、导热垫、绝缘膜和公差链。",
            "用真实 package height、热结构和壳体 CAD 替换目标值。",
        ],
        [
            "3",
            "nRF54L15/nPM1300/NDP120/BMI270 同在 AON PCB 是否合理",
            "PASS WITH CONDITIONS",
            "这四个都是 always-on/low-power domain，放同一 AON/Power PCB 架构合理，能让 nRF 控制主 SoC 电源并保持 BLE/IMU/语音唤醒。",
            "必须隔离 nPM charger 热、nRF RF 匹配、NDP audio clock/decoupling、BMI270 机械振动；NDP/nPM 仍 HOLD。",
        ],
        [
            "4",
            "AON PCB 46 x 14 mm 是否包含完整外围",
            "HOLD",
            "当前是 placement envelope，不是证明完整外围已经能放下；nRF 匹配、nPM 电感/电容、NDP 去耦/Flash、连接器、测试点、保护/电流检测都未完成占位核算。",
            "做 no-route KiCad floorplan 或 ECAD placement envelope。",
        ],
        [
            "5",
            "Compute PCB 62 x 18 mm 是否容纳全部右镜腿电路",
            "HOLD",
            "矩形能摆下不等于能布线。RK3576/LPDDR/RK806/eMMC/TPS61088/FCU760K/MAX98360A/晶振/去耦/测试点/热结构需要 HDI 逃线、DDR 约束和完整高度验证。",
            "8-10 layer HDI 预评审，DDR escape 和热/boost/current loop review。",
        ],
        [
            "6",
            "RK3576 与 LPDDR 逃线和层数是否现实",
            "NEEDS SUPPLIER DATA",
            "只有 RK3576 body 和 LPDDR TBD，缺 ball map、DDR topology、stack-up、via-in-pad/microvia 规则和 length report。",
            "获取 RK3576 HDG/ball map、LPDDR MPN/ball map；完成 DDR pre-layout review。",
        ],
        [
            "7",
            "是否存在电池叠在 RK3576/PMIC/Boost 上方",
            "PASS WITH CONDITIONS",
            "当前 floorplan 把电池纵向错开，右侧 section 明确 no battery under RK3576；这是正确方向。",
            "真实 CAD 中必须把电池膨胀包络、泡棉、FPC 和热片加入，任何热源/电池真实重叠为 FAIL。",
        ],
        [
            "8",
            "Wi-Fi/BLE antenna keep-out 是否冲突",
            "HOLD",
            "当前图保留尾部 keep-out，但天线 SKU、keep-out 尺寸、塑胶窗口、螺丝/铜/扬声器磁体位置都未冻结。",
            "完成 G14 RF/worn-state tuning；speaker magnet 和 battery clearance 同步审查。",
        ],
        [
            "9",
            "左右各一颗电池跨镜腿 1S2P 是否安全",
            "HOLD",
            "架构可行但风险最高：需要同批/容量/DCIR/OCV 匹配、支路 fuse、NTC、branch shunt、FPC 电流和单支路断开行为验证。",
            "关闭 G07；执行 1S2P safety review 和供应商 pack 资格验证。",
        ],
        [
            "10",
            "J4 是否能承受峰值和故障电流",
            "NEEDS SUPPLIER DATA",
            "J4 仍是 custom hinge/interconnect placeholder。pin 数、铜厚、线宽、接触电阻、温升、fault current、bend life 都未冻结。",
            "冻结 FPC stack-up/pinout/current rating；高电流建议多 pin 并联并加支路保护。",
        ],
    ]
    text = [
        "# V2.3 Phase 1 Design Review",
        "",
        f"Generated: {TODAY}",
        "",
        "结论：当前 schematic 是 ERC-clean 的 pre-layout functional schematic，但机械/PCB layout 仍为 HOLD。现在图里的 IC 画在 PCB 上是允许的：它表示器件贴装在同一块 PCB 上，不是实体重叠；真正不允许的是 3D 包络、courtyard、电池膨胀、speaker cavity、RF keep-out 在同一 Z 高度互相冲突。",
        "",
        md_table(["#", "检查项", "结论", "原因", "关闭条件"], rows),
        "",
        "## Immediate Gate Verdict",
        "",
        "- Formal PCB placement/routing: **HOLD**.",
        "- Left temple side/section: **missing**.",
        "- Right temple side/section: **partial evidence only**.",
        "- AON same-board architecture: **allowed with conditions**.",
        "- Compute board 62 x 18 mm: **not proven routable**.",
        "- Cross-temple 1S2P and J4 current path: **major system risk**.",
    ]
    path.write_text("\n".join(text) + "\n", encoding="utf-8")
    return path


OPEN_ITEMS = [
    ("RK806S-5 package / height", "TBD", "exact package/height missing", "Rockchip/Radxa PMIC drawing", "EE", "请提供 RK806S-5 封装图、推荐 land pattern、最大高度、热焊盘和参考电源时序。", "Blocks PMIC placement/height/rail layout", "Official package drawing + land pattern + rail sequence reviewed", "P0"),
    ("LPDDR4X exact MPN / size / ball map", "TBD", "memory MPN not selected", "LPDDR vendor datasheet + ball map", "EE/Sourcing", "请给出 Radxa/Rockchip 已验证 4GB LPDDR4X AVL 料号、封装和 ball map。", "Blocks DDR escape and board layer proof", "MPN + package + ball map + DDR topology + length report", "P0"),
    ("eMMC exact MPN / package", "TBD", "eMMC MPN not selected", "eMMC datasheet/package drawing", "EE/Sourcing", "请给出 32GB eMMC 5.1 AVL 料号、封装、VCC/VCCQ 和 boot validation。", "Blocks boot/storage layout", "MPN + package + BSP boot validation", "P0"),
    ("IMX415 lens / PCB / FPC / total Z", "Current schematic custom module; V2.3 PDF proposes Radxa AS001 32.5 x 32.5 x 18.0", "architecture mismatch", "final module drawing and pinout", "EE/CAD/Supplier", "当前是否接受 32mm Radxa AS001 作为 EVT 机械基线，还是坚持 custom small FPC module？", "Blocks front frame CAD and J3 pinout", "Approved module drawing + lens/FPC/mounting + schematic/BOM update if changed", "P0"),
    ("Wi-Fi antenna", "TBD", "antenna SKU/keep-out not frozen", "antenna vendor drawing + tune report", "RF/CAD", "请提供天线尺寸、塑胶窗口、no-copper keep-out、同轴/FPC 出线和佩戴状态调谐要求。", "Blocks G14 and rear temple layout", "Worn-state antenna tune + keep-out in CAD/PCB", "P0"),
    ("BLE antenna", "TBD", "antenna SKU/keep-out not frozen", "antenna vendor drawing + tune report", "RF/CAD", "请提供 BLE 天线尺寸、净空和 nRF 匹配网络要求。", "Blocks left rear layout", "Antenna SKU + keep-out + matching + tune report", "P0"),
    ("main speaker", "TBD", "MPN/cavity/magnet unknown", "speaker datasheet + acoustic cavity drawing", "Acoustic/CAD", "请提供 8ohm 0.5-1W speaker 的长宽高、磁体、出音孔、声腔和防水网要求。", "Blocks right rear speaker/RF/battery layout", "Speaker MPN + cavity + antenna magnet clearance passed", "P0"),
    ("optional second speaker", "DNP", "not needed in EVT-A but may consume volume", "speaker decision", "PM/Acoustic", "EVT-A 是否删除左侧第二 speaker 机械预留？", "May block left battery/BLE antenna if kept", "DNP exclusion confirmed or cavity allocated", "P1"),
    ("vibration motor", "DNP", "DNP, no MPN", "motor datasheet", "PM/ME", "EVT-A 是否完全删除马达预留？", "Can disturb IMU/mic and consume rear volume", "DNP exclusion confirmed or MPN/keep-out supplied", "P1"),
    ("USB-C connector", "EVT debug tail candidate", "exact connector missing", "USB-C connector drawing", "EE/ME", "请冻结 EVT debug tail Type-C mid-mount MPN 或确认量产只用 pogo。", "Blocks debug tail mechanical opening", "MPN + height + shell opening", "P1"),
    ("magnetic pogo connector", "target about 10 x 3 mm", "target only", "pogo supplier drawing", "EE/ME", "请提供 4-6pin pogo 的针距、高度、电流、接触电阻、磁极、防汗腐蚀和寿命。", "Blocks charge/service rear layout and safety", "MPN + pinout + current/fault/ESD validation", "P0"),
    ("hinge FPC", "6-10 mm wide, 0.10-0.18 mm thick target", "custom stack-up/pinout unknown", "FPC vendor stack-up and impedance report", "EE/ME/FPC", "请提供 fixed-temple EVT cable/FPC 方案，或 folding hinge FPC 的 pin count、铜厚、弯折半径和寿命。", "Blocks J4 current/signal integrity and hinge CAD", "Pinout + current rating + impedance + bend test", "P0"),
    ("Boost inductor", "<=5 x 5 x 2 mm target", "MPN/Isat/DCR/thermal unknown", "inductor datasheet", "EE", "请按 TPS61088、RK3576 boot/AI 峰值电流选择 L1，提供 Isat/DCR/温升。", "Blocks SOC_5V boost layout and thermal", "Selected MPN passes peak/droop/thermal", "P0"),
    ("FH26W exact pin count", "35 pin example only", "final camera/audio/power split unknown", "Hirose drawing + interconnect matrix", "EE", "请根据 MIPI lane、PDM、I2C、GPIO、电源/GND 确认 pin count 和 contact side。", "Blocks J3 footprint and front FPC", "Pin count/contact orientation + impedance checked", "P0"),
    ("RF connector exact MPN", "U.FL/I-PEX class", "exact MPN/height/mating unknown", "Hirose/I-PEX drawing", "RF/EE", "EVT 是否保留 RF debug connector？若保留，请冻结 MPN 和高度。", "Can collide with shell/antenna window", "MPN + height + mating-life approved", "P1"),
    ("system button", "low-profile button/test pad", "exact mechanical actuator missing", "button datasheet / ID decision", "EE/ME", "Power/Recovery/MaskROM 是否做真实侧按键，还是全部隐藏为 pogo test pads？", "Blocks side wall openings and sealing", "MPN/opening/debounce/service plan", "P1"),
    ("battery pack drawing", "70 x 12.8 x 5.6 mm envelope", "supplier max/tolerance/swell not official", "cell/pack drawing + certs", "Battery supplier", "请提供最大尺寸、公差、极耳/线材/PCM/NTC位置、膨胀、IR、倍率和认证。", "Blocks battery bay and 1S2P safety", "Supplier drawing + dummy/real-cell fit test", "P0"),
    ("battery tab and cable exit", "TBD", "tab/cable direction unknown", "battery pack drawing", "Battery supplier/CAD", "请明确极耳/出线方向和最小弯折/应力释放。", "Can collide with speaker/FPC/pogo", "Pack drawing + strain-relief CAD", "P0"),
    ("battery NTC", "10k NTC TBD curve", "B-value/package/placement unknown", "NTC datasheet + charger config", "EE/Battery", "请确认 10k NTC B 值、封装/探头形式、贴附位置和 nPM1300 JEITA 配置。", "Blocks charge safety", "Curve + nPM1300 config + per-cell placement", "P0"),
    ("battery protection", "pack PCM baseline or BQ2970 DNP fallback", "single protection scheme not frozen", "pack PCM spec or BQ2970 suffix/MOSFET", "EE/Battery", "请确认用供应商 PCM 还是板上 BQ2970+MOSFET，不允许两套冲突。", "Blocks pack fault behavior and BOM", "One protection scheme selected and tested", "P0"),
    ("all connector mating heights", "mixed TBD", "J1/J3/J4/J5/USB heights incomplete", "connector drawings", "EE/ME", "请提供所有连接器 mating height、stiffener/FPC厚度和装配方向。", "Blocks Z-height and shell closure", "All connector 3D envelopes in CAD", "P0"),
]


def write_open_items() -> tuple[Path, Path, Path]:
    md = REPORTS / "mechanical_open_items.md"
    csv_path = REPORTS / "mechanical_freeze_matrix.csv"
    xlsx = REPORTS / "mechanical_freeze_matrix.xlsx"
    headers = [
        "Item",
        "Current value",
        "Why it is not frozen",
        "Missing document",
        "Responsible party",
        "Supplier question",
        "Blocking effect",
        "Closure condition",
        "Deadline priority",
    ]
    lines = [
        "# V2.3 Mechanical Open Items",
        "",
        f"Generated: {TODAY}",
        "",
        "所有条目遵守同一规则：没有 exact MPN、package drawing、pinout、land pattern 和最大高度时，不允许进入 release-ready footprint 或正式 layout。",
        "",
        md_table(headers, [list(r) for r in OPEN_ITEMS]),
        "",
    ]
    md.write_text("\n".join(lines), encoding="utf-8")
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(OPEN_ITEMS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mechanical Freeze Matrix"
    ws.append(headers)
    for row in OPEN_ITEMS:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(42, max(14, len(headers[col - 1]) + 4))
    ws.freeze_panes = "A2"
    wb.save(xlsx)
    return md, csv_path, xlsx


def write_footprint_audit(records: list[dict[str, Any]]) -> Path:
    path = REPORTS / "footprint_audit.csv"
    headers = [
        "ref",
        "value",
        "assembly",
        "package_register",
        "freeze_status",
        "body_size",
        "footprint_status",
        "not_for_fab_required",
        "next_action",
    ]
    rows = []
    comps = comp_by_ref()
    for r in records:
        ref = r["reference_designator"]
        c = comps.get(ref)
        value = r["exact_mpn"] if c is None else c.value
        body = f"{r['body_length_mm']} x {r['body_width_mm']} x {r['body_height_mm']}"
        status = r["footprint_status"]
        not_for_fab = "YES" if ("VERIFY" in str(r["land_pattern"]) or r["freeze_status"] in {"HOLD", "TBD"} or r["assembly_status"] == "DNP") else "CHECK"
        rows.append([ref, value, r["assembly_status"], r["land_pattern"], r["freeze_status"], body, status, not_for_fab, r["next_action"]])
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    return path


def write_symbol_to_footprint_mapping(records: list[dict[str, Any]]) -> Path:
    path = REPORTS / "symbol_to_footprint_mapping.csv"
    headers = [
        "ref",
        "bom_id",
        "value",
        "description",
        "board_region",
        "assembly_status",
        "schematic_symbol_strategy",
        "footprint_field",
        "mechanical_package",
        "body_size_mm",
        "freeze_status",
        "footprint_status",
        "not_for_fab_required",
        "3d_model_status",
        "source_document",
        "source_page",
        "gate_or_next_action",
    ]
    by_ref = {r["reference_designator"]: r for r in records}
    rows = []
    for c in COMPONENTS:
        r = by_ref[c.ref]
        not_for_fab = "YES" if ("VERIFY" in c.footprint or r["freeze_status"] in {"HOLD", "TBD"} or c.assembly == "DNP") else "CHECK"
        rows.append(
            [
                c.ref,
                c.bom_id,
                c.value,
                c.desc,
                c.board,
                c.assembly,
                "generated functional block symbol from chipdown_bom.py; not a vendor pin-level symbol unless gate says frozen",
                c.footprint,
                r["package"],
                dims_text(r),
                r["freeze_status"],
                r["footprint_status"],
                not_for_fab,
                r["3d_model_status"],
                r["source_document"],
                r["source_page"],
                c.gate or r["next_action"],
            ]
        )
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    return path


def endpoint_cells(net: str) -> tuple[str, str, str]:
    eps = net_endpoints(net)
    boards = sorted({ep["board"] for ep in eps})
    refs = "; ".join(f"{ep['ref']}.{ep['pin']}={ep['value']}" for ep in eps)
    board_text = "; ".join(boards)
    return refs, board_text, str(len(eps))


def write_interconnect_matrix() -> tuple[Path, Path]:
    csv_path = REPORTS / "interconnect_matrix.csv"
    md_path = REPORTS / "interconnect_matrix.md"
    headers = [
        "net",
        "signal_name",
        "kind",
        "domain",
        "direction",
        "source",
        "destination",
        "default_state",
        "pull",
        "diff_pair",
        "cross_domain",
        "isolation",
        "layout_rule",
        "esd",
        "endpoint_count",
        "boards",
        "endpoints",
        "note",
    ]
    rows = []
    for net in all_nets():
        meta = net_meta(net)
        endpoints, boards, count = endpoint_cells(net)
        rows.append(
            [
                net,
                meta["name"],
                meta["kind"],
                meta["dom"],
                meta["direction"],
                meta["source"],
                meta["dest"],
                meta["default"],
                meta["pull"],
                meta["diff_pair"],
                meta["xdom"],
                meta["isolation"],
                meta["layout"],
                meta["esd"],
                count,
                boards,
                endpoints,
                meta["note"],
            ]
        )
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    crossing_rows = [row for row in rows if row[10] == "yes"]
    high_speed_rows = [row for row in rows if row[2] == "diff" or "ANT" in row[0] or row[0].startswith("CSI_")]
    fpc_rows = [row for row in rows if any(marker in row[15] for marker in ("Front Sensor Board", "R-Temple Compute Board", "L-Temple AON/Power Board")) and row[10] == "yes"]
    lines = [
        "# V2.3 Interconnect Matrix",
        "",
        f"Generated: {TODAY}",
        "",
        "Source of truth: `chipdown_bom.py` component pin lists normalized by `net_meta()`. This is schematic-aligned audit evidence, not a released harness or FPC fabrication drawing.",
        "",
        "## Summary",
        "",
        md_table(
            ["Metric", "Count"],
            [
                ["total schematic nets", len(rows)],
                ["cross-domain / cross-board signals", len(crossing_rows)],
                ["high-speed / RF focus nets", len(high_speed_rows)],
            ],
        ),
        "",
        "## Cross-Domain Signals",
        "",
        md_table(
            ["Net", "Kind", "Domain", "Direction", "Boards", "Isolation / close condition"],
            [[r[0], r[2], r[3], r[4], r[15], r[11]] for r in crossing_rows],
        ),
        "",
        "## High-Speed And RF Nets",
        "",
        md_table(
            ["Net", "Kind", "Diff mate", "Boards", "Layout rule", "ESD"],
            [[r[0], r[2], r[9], r[15], r[12], r[13]] for r in high_speed_rows],
        ),
        "",
        "## FPC / Board-Crossing Focus",
        "",
        md_table(
            ["Net", "Direction", "Boards", "Endpoints"],
            [[r[0], r[4], r[15], r[16]] for r in fpc_rows],
        ),
        "",
        "Full endpoint-level matrix is exported as `v2_chipdown/reports/interconnect_matrix.csv`.",
    ]
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return csv_path, md_path


def write_phase_and_status(records: list[dict[str, Any]]) -> list[Path]:
    counts = status_counts(records)
    phase1 = PHASE / "phase_1_audit.md"
    PHASE.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Phase 1 Audit Report",
        "",
        f"Generated: {TODAY}",
        "",
        "Scope: audit only. No KiCad schematic, symbol, footprint, PCB or 3D model files were modified by this phase.",
        "",
        "## Evidence Reviewed",
        "",
        "- Current schematic/BOM source: `v2_chipdown/scripts/chipdown_bom.py`.",
        "- Current KiCad schematic: `v2_chipdown/hardware/ai_glasses_v2_chipdown.kicad_sch`.",
        "- Existing ERC report: `v2_chipdown/reports/erc.report.txt` shows 0 errors / 0 warnings on 2026-07-01T01:26:33.",
        "- Existing Phase 1.5 floorplan render/report.",
        "- Symbol/footprint audit: `v2_chipdown/reports/symbol_to_footprint_mapping.csv`.",
        "- Interconnect audit: `v2_chipdown/reports/interconnect_matrix.csv` and `.md`.",
        f"- Attached dimensions PDFs: `{DIM_PDF_V22}` and `{DIM_PDF_V23}`.",
        "- Local HOLD closure pack under `AI_Glasses_HOLD_Closure_Pack/`.",
        "",
        "## Database Status Counts",
        "",
        md_table(["Status", "Count"], [[k, v] for k, v in sorted(counts.items())]),
        "",
        "## Phase 1 Verdict",
        "",
        "- Schematic is useful for review and procurement, but PCB layout is **HOLD**.",
        "- Major blockers: left side/sections missing, compute routability unproven, 1S2P/J4 safety not proven, RF/speaker/battery/camera mechanical data incomplete.",
        "- The external V2.3 dimensions PDF camera baseline is not fully aligned with the current schematic because U14 is still a custom IMX415 FPC module in HOLD.",
        "",
        "## Next Phase Gate",
        "",
        "Do not enter Phase 2 freeze or Phase 4 KiCad modification until supplier documents are gathered for the P0 open items, especially RK806S, LPDDR4X, eMMC, battery pack, J3/J4, antennas, speaker and camera module.",
    ]
    phase1.write_text("\n".join(lines) + "\n", encoding="utf-8")

    p0_items = [item for item in OPEN_ITEMS if item[-1] == "P0"]
    local_sources = [
        ["RK3576", SOURCES["rk3576"].title, "HOLD", "brief datasheet only; HDG/ball map/land pattern missing"],
        ["FCU760K", SOURCES["fcu760k"].title, "HOLD", "outline and basic electrical data present; hardware design guide/land pattern missing"],
        ["NDP120", SOURCES["ndp120"].title, "HOLD", "product brief only; full datasheet/rail map/package drawing missing"],
        ["IMX415", SOURCES["imx415"].title, "HOLD", "bare sensor source present; final custom module/lens/FPC drawing missing"],
        ["BQ2970", SOURCES["bq2970"].title, "DNP", "valid fallback IC source; pack PCM scheme still must be chosen"],
        ["LP451165", SOURCES["battery_fit"].title, "HOLD", "project fit envelope present; supplier max drawing/certs/tabs/NTC missing"],
    ]
    phase2 = PHASE / "phase_2_supplier_data_freeze.md"
    phase2.write_text(
        "\n".join(
            [
                "# Phase 2 Supplier Data Freeze Status",
                "",
                f"Generated: {TODAY}",
                "",
                "Verdict: **HOLD**. Local source documents were reviewed and folded into the mechanical database, but the supplier-data set is not sufficient to freeze layout or release footprints.",
                "",
                "## Local Source Review",
                "",
                md_table(["Item", "Source", "Freeze status", "Reason"], local_sources),
                "",
                "## P0 Supplier Questions Blocking PCB Layout",
                "",
                md_table(
                    ["Item", "Missing document", "Closure condition"],
                    [[item[0], item[3], item[7]] for item in p0_items],
                ),
                "",
                "No online supplier data was promoted to frozen status in this phase. Per the project rule, only official datasheets, package drawings, land patterns and verified CAD/STEP files can close these items.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    phase3 = PHASE / "phase_3_database_pdf_report.md"
    phase3.write_text(
        "\n".join(
            [
                "# Phase 3 Database And PDF Report",
                "",
                f"Generated: {TODAY}",
                "",
                "Verdict: **PASS WITH CONDITIONS** for documentation generation, **HOLD** for mechanical freeze.",
                "",
                "The single source used by reports is `v2_chipdown/config/hardware_mechanical_database.yaml`, generated from the current schematic/BOM plus conservative source overlays.",
                "",
                "## Generated Database / Report Artifacts",
                "",
                md_table(
                    ["Artifact", "Purpose"],
                    [
                        ["hardware_mechanical_database.yaml/csv", "source-bound dimensions, power notes, statuses and next actions"],
                        ["AI_Glasses_Hardware_Dimensions_and_Performance_V2.3.md/pdf", "human-readable hardware dimension/performance report with component cards"],
                        ["AI_Glasses_V2_3_Mechanical_Views.svg/png/pdf", "16 required schematic-aligned mechanical views"],
                        ["mechanical_freeze_matrix.xlsx/csv", "P0/P1 open items and closure conditions"],
                    ],
                ),
                "",
                "Current status counts:",
                "",
                md_table(["Status", "Count"], [[k, v] for k, v in sorted(counts.items())]),
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    phase4 = PHASE / "phase_4_kicad_update_decision.md"
    phase4.write_text(
        "\n".join(
            [
                "# Phase 4 KiCad Update Decision",
                "",
                f"Generated: {TODAY}",
                "",
                "Verdict: **HOLD / NO KiCad EDITS IN THIS RUN**.",
                "",
                "Reason: the prompt explicitly forbids release-ready footprint or schematic changes when exact MPN, package drawing, pinout, land pattern and max height are missing. The current project still has P0 blockers for RK806S, LPDDR4X, eMMC, camera module, antennas, speaker, pogo, J3/J4, boost inductor and battery pack drawing.",
                "",
                "## KiCad Audit Evidence Generated",
                "",
                md_table(
                    ["Artifact", "Result"],
                    [
                        ["symbol_to_footprint_mapping.csv", f"{len(COMPONENTS)} schematic components mapped to current footprint fields and release-readiness status"],
                        ["footprint_audit.csv", "NOT_FOR_FAB requirements flagged for VERIFY/HOLD/TBD/DNP items"],
                        ["interconnect_matrix.csv/md", f"{len(all_nets())} schematic nets with endpoints, domains, crossing status and layout rules"],
                        ["schematic_diff.md", "confirms no KiCad schematic edits were made"],
                    ],
                ),
                "",
                "Allowed future KiCad action: create HOLD placeholders or a no-route mechanical floorplan only after the team approves that output as non-fabrication evidence. Do not replace current functional symbols with vendor pin-level symbols until ball maps/pinouts are official.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    phase5 = PHASE / "phase_5_release_gate.md"
    phase5.write_text(
        "\n".join(
            [
                "# Phase 5 ERC / DRC / 3D / Release Gate",
                "",
                f"Generated: {TODAY}",
                "",
                "Overall verdict: **NOT READY FOR PCB LAYOUT**.",
                "",
                "## Electrical / Layout Checks",
                "",
                md_table(
                    ["Check", "Status", "Evidence / blocker"],
                    [
                        ["ERC", "PASS on existing report", "`v2_chipdown/reports/erc.report.txt` shows 0 errors / 0 warnings on 2026-07-01T01:26:33; `kicad-cli` unavailable here, so not rerun"],
                        ["DRC", "NOT RUN", "V2 has no `.kicad_pcb`; DRC requires at least a no-route or formal PCB file"],
                        ["3D collision", "NOT RUN", "No verified STEP/3D model set aligned to V2 KiCad footprints"],
                        ["Thermal", "HOLD", "RK3576/RK806/TPS61088 heat path and battery isolation require CAD/thermal review"],
                        ["RF", "HOLD", "Wi-Fi/BLE antenna SKU, keep-out and worn-state tune missing"],
                        ["Battery/J4 safety", "HOLD", "1S2P branch protection/current capacity/fault behavior not frozen"],
                    ],
                ),
                "",
                "PCB layout may start only as a controlled floorplan study. Formal routing/fab release remains blocked until all P0 closure conditions in `mechanical_open_items.md` are closed.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    drc = REPORTS / "drc_report.txt"
    drc.write_text(
        "\n".join(
            [
                f"DRC report generated by Phase 1 audit on {TODAY}",
                "",
                "NOT RUN: V2 chipdown project currently has no .kicad_pcb file.",
                "This is intentional: formal PCB placement/routing is blocked by Gate 0 / G00F.",
                "A real DRC can only be run after the no-route floorplan or formal PCB file exists.",
                "",
                "Current available electrical evidence:",
                "- v2_chipdown/reports/erc.report.txt: 0 errors / 0 warnings on 2026-07-01T01:26:33.",
                "- kicad-cli was not available in this shell session, so ERC was not rerun in this phase.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    changed = REPORTS / "changed_files.md"
    changed.write_text(
        "\n".join(
            [
                "# Phase 1 Changed Files",
                "",
                "Audit-generated files only. No KiCad schematic/footprint/PCB edits were made in Phase 1.",
                "",
                "- `v2_chipdown/config/hardware_mechanical_database.yaml`",
                "- `v2_chipdown/config/hardware_mechanical_database.csv`",
                "- `v2_chipdown/reports/design_review.md`",
                "- `v2_chipdown/reports/mechanical_open_items.md`",
                "- `v2_chipdown/reports/mechanical_freeze_matrix.csv`",
                "- `v2_chipdown/reports/mechanical_freeze_matrix.xlsx`",
                "- `v2_chipdown/reports/footprint_audit.csv`",
                "- `v2_chipdown/reports/symbol_to_footprint_mapping.csv`",
                "- `v2_chipdown/reports/interconnect_matrix.csv`",
                "- `v2_chipdown/reports/interconnect_matrix.md`",
                "- `v2_chipdown/reports/phase_reports/phase_1_audit.md`",
                "- `v2_chipdown/reports/phase_reports/phase_2_supplier_data_freeze.md`",
                "- `v2_chipdown/reports/phase_reports/phase_3_database_pdf_report.md`",
                "- `v2_chipdown/reports/phase_reports/phase_4_kicad_update_decision.md`",
                "- `v2_chipdown/reports/phase_reports/phase_5_release_gate.md`",
                "- `v2_chipdown/reports/drc_report.txt`",
                "- `v2_chipdown/reports/schematic_diff.md`",
                "- `v2_chipdown/reports/engineering_reviews/*.md`",
                "- `v2_chipdown/reports/output/mechanical_freeze/*`",
                "- `v2_chipdown/reports/AI_Glasses_Hardware_Dimensions_and_Performance_V2.3.md`",
                "- `v2_chipdown/reports/output/AI_Glasses_Hardware_Dimensions_and_Performance_V2.3.pdf`",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    diff = REPORTS / "schematic_diff.md"
    diff.write_text(
        "\n".join(
            [
                "# Phase 1 Schematic Diff",
                "",
                "No KiCad schematic edits were made in this audit phase.",
                "",
                "Required future schematic decisions before Phase 4:",
                "",
                "- Decide whether U14 stays `IMX415-AAQR module (custom FPC)` or changes to Radxa Camera 4K AS001 mechanical baseline.",
                "- Add/verify any NOT_FOR_FAB markings for VERIFY/HOLD/TBD packages once actual KiCad footprint libraries exist.",
                "- Freeze J3/J4 pin counts and update interconnect matrix before any PCB placement.",
                "- Freeze RK806S, LPDDR4X, eMMC and FCU760K official package/land patterns before assigning release footprints.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    return [phase1, phase2, phase3, phase4, phase5, drc, changed, diff]


def write_engineering_reviews() -> list[Path]:
    REVIEWS.mkdir(parents=True, exist_ok=True)
    reviews = {
        "1S2P_cross_temple_safety_review.md": [
            "# 1S2P Cross-Temple Safety Review",
            "",
            "Verdict: **HOLD**.",
            "",
            "- Cell matching: require same lot, same MPN, OCV window, capacity window and DCIR window before parallel connection.",
            "- Equalization current: must be limited by branch resistance/fuse/PTC and verified before connecting cells across the frame.",
            "- Fuse/protection: F1/F2 branch protection and one chosen pack protection scheme are mandatory; do not populate incompatible double-protection.",
            "- NTC: one NTC per cell, thermally bonded to the cell body and read by nPM1300/AON ADC.",
            "- FPC current capacity: J4/FPC copper width, thickness, pin count and temperature rise are not frozen.",
            "- Connector resistance: contact resistance must be in the voltage-drop and heating budget.",
            "- Voltage drop: RS1/RS2/RS3 plus FPC/contact resistance must not trip boost UVLO during RK3576 boot/AI bursts.",
            "- Short-circuit path: worst case is one cell/FPC branch shorting into the other cell. Branch fuse/PTC must open before FPC damage.",
            "- Single-cell disconnect: system must detect imbalance or branch loss and derate/stop charging.",
        ],
        "power_tree_and_sequence.md": [
            "# Power Tree And Sequence Review",
            "",
            "Verdict: **PASS WITH CONDITIONS** for architecture, **HOLD** for timing/measurement.",
            "",
            "- Always-on power: BT1/BT2 -> branch protection/shunts -> BAT_P -> RS1 -> nPM1300 -> AON_1V8/AON_3V3.",
            "- Main boost: nRF54L15 asserts SOC_PWR_EN, TPS61088 generates SOC_5V, SOC_5V_PGOOD returns to AON.",
            "- RK3576 rails: RK806S generates CPU/GPU/NPU/logic/DDR/IO rails; exact sequencing requires RK806S/RK3576 HDG.",
            "- Wake sequence: BLE/IMU/NDP/button event -> nRF -> enable boost -> RK806 PWRON -> RK3576 alive/status.",
            "- Shutdown: RK3576 requests/acknowledges safe off -> nRF cuts camera/Wi-Fi/audio then SOC_5V.",
            "- Brownout: boost UVLO and nPM1300 thresholds must be coordinated with battery sag and branch-current imbalance.",
        ],
        "pcb_routability_review.md": [
            "# PCB Routability Review",
            "",
            "Verdict: **HOLD**.",
            "",
            "- Estimated layer count: right compute island likely 8-10 layer HDI, not a simple 4-layer wearable board.",
            "- DDR escape: unproven until RK3576/LPDDR ball maps and selected LPDDR MPN exist.",
            "- BGA via technology: expect microvia/via-in-pad or very tight fanout rules; must be quoted by PCB fab early.",
            "- Controlled impedance: MIPI CSI, USB2, Wi-Fi USB and any RF/coax transitions need stack-up from layout/fab.",
            "- High-speed corridor: RK3576-LPDDR, RK3576-eMMC and front FPC corridors must be reserved before placing battery/speaker.",
            "- Power planes: SOC_5V, RK806 rails and GND return must not be fragmented by cosmetic temple shape.",
        ],
        "thermal_review.md": [
            "# Thermal Review",
            "",
            "Verdict: **HOLD**.",
            "",
            "- RK3576: heat must go to the outer temple wall/heat spreader, not skin side and not into the cell.",
            "- RK806S: PMIC heat and inductor/cap placement need thermal current budget.",
            "- TPS61088/boost: high-current loop and L1 can become local hot spots.",
            "- Wi-Fi: FCU760K transmit bursts add local heat and RF keep-out constraints.",
            "- Amplifier/speaker: peak audio power is acoustic/thermal dependent; speaker magnet must clear antenna.",
            "- Battery isolation: no hot component or heat spreader should directly press or radiate into LP451165 envelope.",
            "- Skin-side path: add NTC at hot spreader and skin side; firmware throttles record/AI burst.",
        ],
        "rf_review.md": [
            "# RF Review",
            "",
            "Verdict: **HOLD**.",
            "",
            "- BLE/Wi-Fi spacing: current concept separates BLE left rear and Wi-Fi right rear, which is good.",
            "- Keep-out: antenna windows must clear battery, copper, screws, speaker magnet, pogo magnets and boost current loops.",
            "- Ground clearance: final FPC/PCB/LDS antenna vendor rules must drive the CAD/PCB no-copper area.",
            "- Battery/metal interference: cells and shields detune antennas; worn-head tuning is mandatory.",
            "- Coexistence: FCU760K shares Wi-Fi/BT antenna in selected ordering code; nRF BLE antenna coexistence requires test plan.",
            "- Test method: tune in shell, with battery, speaker, pogo, glasses on phantom/head fixture and active Wi-Fi/BLE traffic.",
        ],
        "mass_and_balance.md": [
            "# Mass And Balance Review",
            "",
            "Verdict: **HOLD**.",
            "",
            "- Left temple mass: BT2 about 6 g estimated plus AON PCB; exact mass TBD.",
            "- Right temple mass: BT1 about 6 g estimated plus RK3576 compute, Wi-Fi and speaker, likely heavier than left.",
            "- Front frame mass: camera/lens choice dominates. Radxa AS001-style 32 mm board is much larger than a custom small FPC module.",
            "- Center of gravity: right front compute pod may pull glasses forward/right unless left mass or frame geometry balances it.",
            "- Left-right imbalance: target should be measured with dummy cells/PCBs/speaker/camera before ID tooling.",
            "- Action: create mass spreadsheet once MPNs, PCBs and shell CAD have weights.",
        ],
        "assembly_tolerance_stack.md": [
            "# Assembly Tolerance Stack Review",
            "",
            "Verdict: **HOLD**.",
            "",
            "- Shell: wall thickness, draft, bosses, snap features and screw locations are not frozen.",
            "- PCB: board thickness, copper stack, component-side allocation and solder height need CAD import.",
            "- Battery: use 70 x 12.8 x 5.6 mm clearance envelope until supplier max/swell data arrives.",
            "- Adhesive/foam: reserve compression range and avoid point loads on pouch cells.",
            "- FPC: bend radius, stiffener, soldered areas and no-via dynamic zone must be defined.",
            "- Connector: mating heights and contact side for J3/J4/J1/J5 must be included.",
            "- Component height: no official package drawing means no Z-height freeze.",
            "- Manufacturing tolerance: final fit must include worst-case shell + PCB + component + battery + adhesive stack.",
        ],
    }
    paths = []
    for filename, lines in reviews.items():
        p = REVIEWS / filename
        p.write_text("\n".join(lines) + "\n", encoding="utf-8")
        paths.append(p)
    return paths


def dims_text(r: dict[str, Any]) -> str:
    vals = [r["body_length_mm"], r["body_width_mm"], r["body_height_mm"]]
    if all(v in (None, "", "TBD") for v in vals):
        return "TBD"
    return " x ".join(str(v) if v not in (None, "") else "TBD" for v in vals) + " mm"


def short(text: str, width: int = 72) -> str:
    return "\n".join(textwrap.wrap(str(text), width=width, replace_whitespace=False)) or ""


def write_performance_report(records: list[dict[str, Any]]) -> Path:
    path = REPORTS / "AI_Glasses_Hardware_Dimensions_and_Performance_V2.3.md"
    major = ordered_major_records(records)
    lines = [
        "# AI Glasses Hardware Dimensions and Performance V2.3",
        "",
        f"Generated: {TODAY}",
        "",
        "This report is generated from the current schematic/BOM plus the single mechanical database. Current schematic wins over older PDFs. Numbers without official package drawings remain HOLD/TBD/PROVISIONAL.",
        "",
        "## Executive Summary",
        "",
        "- Normal thin optical frames do not fit this RK3576 chip-down architecture.",
        "- Thick smart-glasses / sports-sunglasses temples can conditionally fit if Phase 1.5 envelopes are met.",
        "- AON and compute boards intentionally contain multiple ICs: they are real PCB assemblies, not overlapping solid objects.",
        "- Current PCB layout is not released. KiCad currently has no V2 `.kicad_pcb`; DRC is not applicable yet.",
        "- The biggest risks before PCB layout are battery/FPC current safety, DDR/HDI routability, antenna keep-out, camera module choice, and Z-height tolerance stack.",
        "",
        "## Current Power States",
        "",
        "| State | Current model | Target / use case |",
        "|---|---:|---|",
        "| Deep Off / AON only | 22 mW | target 20-50 mW |",
        "| Quick Standby | 134 mW | DDR/light standby model |",
        "| Phone-assisted safety ID | 362 mW | 3-5 h blended target |",
        "| Mixed motion use | 450 mW | target <=500-800 mW |",
        "| Continuous 1080p record | 1290 mW | 45-90 min product target |",
        "| AI + record burst | 2065 mW | 30-60 min / short burst target |",
        "",
        "## Frozen / Open Dimension Summary",
        "",
        md_table(
            ["Ref", "MPN / item", "Board", "Dims", "Status", "Source", "Next action"],
            [
                [
                    r["reference_designator"],
                    r["exact_mpn"],
                    r["board_region"],
                    dims_text(r),
                    r["freeze_status"],
                    r["source_document"],
                    r["next_action"],
                ]
                for r in major
            ],
        ),
        "",
        "## Component Cards",
        "",
    ]
    for r in major:
        lines.extend(
            [
                f"### {r['reference_designator']} - {r['exact_mpn']}",
                "",
                f"- What it is: {r['description'] or r['functional_block']}.",
                f"- What it does: {plain_function(r)}",
                f"- Why needed: {plain_why(r)}",
                f"- Connects to: {plain_connects(r)}",
                f"- Voltage: {r['operating_voltage']}.",
                f"- Power: typical {r['typical_power']}; peak {r['peak_power']}.",
                f"- Body size: {dims_text(r)}.",
                f"- Extra area: {plain_extra_area(r)}",
                f"- Placement: {r['board_region']}; {r['placement_notes']}",
                f"- Keep away from: {plain_keepaway(r)}",
                f"- Freeze status: {r['freeze_status']}; footprint: {r['footprint_status']}.",
                f"- Missing data: {r['next_action']}",
                f"- Source: {r['source_document']} ({r['source_revision']}), {r['source_page']}, `{r['source_url_or_file']}`.",
                "",
            ]
        )
    lines.extend(
        [
            "## Source Register",
            "",
            md_table(
                ["ID", "Title", "Revision", "Page", "Local file / URL", "Confidence"],
                [[v.doc, v.title, v.revision, v.page, v.url_or_file, v.confidence] for v in SOURCES.values()],
            ),
            "",
            "## Release Checklist",
            "",
            "- Left/right temple top, side and section views: generated as Phase 1 evidence; still HOLD until CAD validates.",
            "- Every major hardware-list item has a component card and freeze status.",
            "- Every dimension has a source register entry or is marked TBD/HOLD.",
            "- KiCad footprint edits: not performed in Phase 1.",
            "- Symbol/footprint mapping and interconnect matrix: generated from current schematic source.",
            "- ERC: existing report clean; not rerun because `kicad-cli` was unavailable.",
            "- DRC/3D collision: not complete because V2 has no PCB/STEP assembly yet.",
            "- Final PASS/HOLD/FAIL: overall **HOLD** before PCB layout.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def plain_function(r: dict[str, Any]) -> str:
    block = r["functional_block"].lower()
    if "battery" in block:
        return "提供整机电源和续航。"
    if "antenna" in block:
        return "把无线芯片的 RF 信号发射/接收出去。"
    if "camera" in block:
        return "采集画面或给相机供电/保护。"
    if "compute" in block or "soc" in block:
        return "运行 Linux/AI/编码/ISP 等高负载任务。"
    if "aon" in block or "mcu" in block:
        return "在主 SoC 断电时保持 BLE、IMU、唤醒和电源控制。"
    if "speaker" in block or "amplifier" in block:
        return "输出提示音或语音。"
    if "fpc" in block or "connector" in block or "pogo" in block:
        return "连接板卡、电源、信号或充电/调试接口。"
    return "完成该功能模块在系统中的专用任务。"


def plain_why(r: dict[str, Any]) -> str:
    ref = r["reference_designator"]
    mapping = {
        "U1": "需要本地 AI、摄像头 ISP 和视频编码时才启动。",
        "U7": "它是低功耗常开大脑，负责让主 SoC 彻底断电。",
        "U9": "让语音唤醒不需要一直开 RK3576。",
        "U10": "检测运动、姿态和佩戴状态。",
        "BT1": "右侧电池提供容量并平衡整机重量。",
        "BT2": "左侧电池与右侧并联形成 1S2P。",
    }
    return mapping.get(ref, "当前架构需要它来完成对应电源、信号、传感、射频、声学或调试功能。")


def plain_connects(r: dict[str, Any]) -> str:
    ref = r["reference_designator"]
    mapping = {
        "U1": "RK806S、LPDDR4X、eMMC、camera MIPI、Wi-Fi USB、audio I2S、AON UART。",
        "U7": "nPM1300、NDP120、BMI270、AON I2C、J4 interconnect 和各路 enable/PGOOD。",
        "U11": "WIFI_3V3、USB2 to RK3576、PCM/audio、Wi-Fi antenna。",
        "U14": "J3 FPC、CAM rails、MIPI CSI、CAM I2C/MCLK/RST/PWDN。",
        "BT1": "F1/RS3 branch path into BAT_P and J4/AON power path。",
        "BT2": "F2/RS2 branch path into BAT_P and nPM1300。",
        "J4": "left AON/power side and right compute/battery side。",
    }
    return mapping.get(ref, "见 current schematic nets and board partition。")


def plain_extra_area(r: dict[str, Any]) -> str:
    block = r["functional_block"].lower()
    if "boost" in block:
        return "还要电感、输入/输出电容、反馈、热铜和高电流回路空间。"
    if "antenna" in block:
        return "还要 no-copper keep-out、塑胶窗口和 matching network。"
    if "battery" in block:
        return "还要极耳/线材、NTC、泡棉、膨胀、绝缘和维修余量。"
    if "camera" in block:
        return "还要镜头、固定孔、FPC、ESD、供电小板和 Z-stack。"
    if "soc" in block or "memory" in block:
        return "还要去耦、fanout、DDR corridor、热扩散和测试点。"
    return "还要去耦、连接器、测试点、keep-out 和装配公差。"


def plain_keepaway(r: dict[str, Any]) -> str:
    block = r["functional_block"].lower()
    if "battery" in block:
        return "RK3576、PMIC、boost、电感、尖锐边、螺丝、热扩散片和强压点。"
    if "antenna" in block:
        return "电池、铜皮、螺丝、speaker 磁体、pogo 磁体、高电流环路。"
    if "imu" in block:
        return "speaker、马达、高应力壳体筋位和剧烈振动源。"
    if "camera" in block:
        return "高温 regulator、电池硬压、过紧 FPC 弯折和污染源。"
    if "boost" in block:
        return "antenna、IMU、mic、skin-side hot spot 和电池。"
    return "天线 keep-out、电池膨胀区、高热源和不可维修的硬压位置。"


def draw_package(ax, r: dict[str, Any]) -> None:
    ax.axis("off")
    ax.set_xlim(0, 100)
    ax.set_ylim(0, 72)
    status = r["freeze_status"]
    color = "#" + STATUS_COLORS.get(status, "D9E1F2")
    ax.add_patch(patches.FancyBboxPatch((1, 1), 98, 70, boxstyle="round,pad=0.8,rounding_size=4", facecolor="#ffffff", edgecolor="#333333", linewidth=0.8))
    ax.add_patch(patches.Rectangle((1, 62), 98, 9, facecolor=color, edgecolor="#333333", linewidth=0.8))
    display_title = f"{r['reference_designator']}  {r['exact_mpn']}"
    if len(display_title) > 52:
        display_title = f"{r['reference_designator']}  {r['functional_block']}"
    ax.text(4, 66.5, display_title, fontsize=7.2, fontweight="bold", va="center")
    ax.text(94, 66.5, status, fontsize=6.8, fontweight="bold", va="center", ha="right")

    # Simplified package figure.
    l = r["body_length_mm"]
    w = r["body_width_mm"]
    h = r["body_height_mm"]
    try:
        lf = float(l)
        wf = float(w)
    except Exception:
        lf = wf = 10.0
    scale = min(42 / max(lf, 1), 22 / max(wf, 1))
    bw = max(10, lf * scale)
    bh = max(7, wf * scale)
    x = 8
    y = 34
    ax.add_patch(patches.Rectangle((x, y), bw, bh, facecolor="#e7f5ff", edgecolor="#111111", linewidth=1.0))
    ax.add_patch(patches.Circle((x + 3, y + bh - 3), 1.2, facecolor="#111111"))
    ax.text(x + bw / 2, y - 3, dims_text(r), fontsize=5.9, ha="center", va="top")
    ax.text(56, 56, short(r["functional_block"], 38), fontsize=6.0, va="top")
    ax.text(56, 43, "Voltage: " + short(r["operating_voltage"], 33), fontsize=5.3, va="top")
    ax.text(56, 30, "Power: " + short(f"{r['typical_power']} / peak {r['peak_power']}", 33), fontsize=5.3, va="top")
    ax.text(5, 23, "Purpose: " + short(plain_function(r), 76), fontsize=5.5, va="top")
    ax.text(5, 14, "Placement: " + short(r["board_region"], 76), fontsize=5.3, va="top")
    ax.text(5, 8, "Missing: " + short(r["next_action"], 76), fontsize=5.1, va="top")


MECH_COLORS = {
    "frame": "#f2f2ef",
    "compute": "#8ecae6",
    "aon": "#b7e4c7",
    "camera": "#ffb4a2",
    "battery": "#ffe066",
    "rf": "#cdb4db",
    "speaker": "#f4a261",
    "fpc": "#adb5bd",
    "pogo": "#90be6d",
    "heat": "#ef476f",
    "shell": "#f8f9fa",
    "text": "#202020",
    "outline": "#242424",
}


def view_box(ax, x: float, y: float, w: float, h: float, label: str, color: str, *, hatch: str | None = None, fontsize: float = 7.0, lw: float = 1.0, alpha: float = 0.94) -> None:
    ax.add_patch(
        patches.Rectangle(
            (x, y),
            w,
            h,
            facecolor=color,
            edgecolor=MECH_COLORS["outline"],
            linewidth=lw,
            hatch=hatch,
            alpha=alpha,
        )
    )
    if label:
        ax.text(x + w / 2, y + h / 2, label, ha="center", va="center", fontsize=fontsize, color=MECH_COLORS["text"], linespacing=1.05)


def view_dim(ax, x0: float, y0: float, x1: float, y1: float, label: str, *, fontsize: float = 6.5) -> None:
    ax.annotate("", xy=(x1, y1), xytext=(x0, y0), arrowprops=dict(arrowstyle="<->", lw=0.75, color="#333333"))
    ax.text((x0 + x1) / 2, (y0 + y1) / 2, label, fontsize=fontsize, ha="center", va="bottom", color="#333333")


def setup_view(ax, title: str, xlim: tuple[float, float], ylim: tuple[float, float], *, grid: bool = True) -> None:
    ax.set_title(title, loc="left", fontsize=8.5, fontweight="bold")
    ax.set_xlim(*xlim)
    ax.set_ylim(*ylim)
    ax.set_aspect("auto")
    if grid:
        ax.grid(True, linewidth=0.25, color="#d0d0d0", alpha=0.65)
        ax.tick_params(labelsize=5.8)
    else:
        ax.axis("off")
    for spine in ax.spines.values():
        spine.set_linewidth(0.8)
        spine.set_color("#333333")


def draw_whole_top_view(ax) -> None:
    setup_view(ax, "1 Whole glasses top view", (-170, 310), (-42, 72), grid=False)
    front = patches.FancyBboxPatch((0, 0), 135, 34, boxstyle="round,pad=0.02,rounding_size=8", facecolor=MECH_COLORS["frame"], edgecolor=MECH_COLORS["outline"], linewidth=1.2)
    ax.add_patch(front)
    ax.add_patch(patches.FancyBboxPatch((8, 5), 52, 24, boxstyle="round,pad=0.02,rounding_size=7", facecolor="#ffffff", edgecolor=MECH_COLORS["outline"], linewidth=0.8))
    ax.add_patch(patches.FancyBboxPatch((75, 5), 52, 24, boxstyle="round,pad=0.02,rounding_size=7", facecolor="#ffffff", edgecolor=MECH_COLORS["outline"], linewidth=0.8))
    view_box(ax, 54, 30, 28, 7, "front sensor PCB\ncamera + mics", MECH_COLORS["camera"], fontsize=5.5)
    for x in (18, 48, 88, 118):
        ax.add_patch(patches.Circle((x, 33), 1.4, facecolor="#222222"))
    view_box(ax, 135, 18, 150, 20, "RIGHT TEMPLE", "#e7f5ff", fontsize=7.2)
    view_box(ax, -145, 18, 145, 16, "LEFT TEMPLE", "#ebfbee", fontsize=7.2)
    view_box(ax, 143, 21, 62, 14, "compute\n62x18", MECH_COLORS["compute"], fontsize=5.8)
    view_box(ax, 209, 21, 70, 12, "BT1\n70x12.8", MECH_COLORS["battery"], fontsize=5.8)
    view_box(ax, 262, 33, 20, 4, "Wi-Fi ant.", MECH_COLORS["rf"], hatch="//", fontsize=5)
    view_box(ax, -136, 20, 46, 12, "AON/power\n46x14", MECH_COLORS["aon"], fontsize=5.8)
    view_box(ax, -84, 20, 70, 12, "BT2\n70x12.8", MECH_COLORS["battery"], fontsize=5.8)
    view_box(ax, -30, 31, 20, 3.5, "BLE ant.", MECH_COLORS["rf"], hatch="//", fontsize=5)
    ax.annotate("J3 front FPC: MIPI CSI, camera rails, PDM/I2C/GND", xy=(132, 33), xytext=(92, 58), arrowprops=dict(arrowstyle="->", lw=0.85), fontsize=6.1)
    ax.annotate("J4 L<->R: BAT_P/VSYS, AON UART, EN/PGOOD, GND", xy=(0, 22), xytext=(-55, -27), arrowprops=dict(arrowstyle="->", lw=0.85), fontsize=6.1)
    ax.annotate("1S2P: one LP451165 per temple", xy=(210, 24), xytext=(152, -27), arrowprops=dict(arrowstyle="->", lw=0.85), fontsize=6.1)


def draw_whole_side_view(ax) -> None:
    setup_view(ax, "2 Whole glasses side view", (-5, 155), (-1, 10), grid=True)
    ax.set_xlabel("length from hinge / mm", fontsize=6)
    ax.set_ylabel("thickness / mm", fontsize=6)
    view_box(ax, 0, 0, 150, 8.5, "", MECH_COLORS["frame"], lw=1.3)
    view_box(ax, 8, 5.2, 62, 1.0, "compute PCB", MECH_COLORS["compute"], fontsize=5.8)
    view_box(ax, 11, 6.2, 22, 1.6, "RK3576 heat", "#48cae4", fontsize=5.5)
    view_box(ax, 74, 1.0, 70, 5.6, "battery 5.6 max + swell/foam", MECH_COLORS["battery"], fontsize=6.2)
    view_box(ax, 108, 6.6, 16, 1.7, "speaker", MECH_COLORS["speaker"], fontsize=5.8)
    ax.plot([8, 70], [8.2, 8.2], color=MECH_COLORS["heat"], linewidth=2.2)
    ax.text(11, 8.55, "heat to outer wall; no battery under RK3576", fontsize=5.8, color=MECH_COLORS["heat"])
    view_dim(ax, 151, 0, 151, 8.5, ">=8.5 mm")


def draw_front_front_view(ax) -> None:
    setup_view(ax, "3 Front frame front view", (0, 100), (0, 55), grid=False)
    ax.add_patch(patches.FancyBboxPatch((9, 12), 82, 30, boxstyle="round,pad=0.02,rounding_size=8", facecolor=MECH_COLORS["frame"], edgecolor=MECH_COLORS["outline"], linewidth=1.1))
    ax.add_patch(patches.Circle((32, 27), 10, fill=False, linewidth=1.1))
    ax.add_patch(patches.Circle((68, 27), 10, fill=False, linewidth=1.1))
    view_box(ax, 42, 40, 16, 5, "U14\nHOLD", MECH_COLORS["camera"], fontsize=5.5)
    for x in (22, 42, 58, 78):
        ax.add_patch(patches.Circle((x, 44), 1.2, facecolor="#222"))
    ax.text(50, 7, "camera + T5837 mic ports; final module drawing still HOLD", fontsize=5.6, ha="center")


def draw_front_top_view(ax) -> None:
    setup_view(ax, "4 Front frame top view", (0, 100), (0, 55), grid=False)
    view_box(ax, 18, 20, 64, 10, "front brow structural channel", MECH_COLORS["frame"], fontsize=5.8)
    view_box(ax, 42, 29, 16, 5, "sensor PCB\n32x8x5 target", MECH_COLORS["camera"], fontsize=5.3)
    view_box(ax, 16, 14, 68, 2.5, "J3 FPC path", MECH_COLORS["fpc"], fontsize=5.3)
    ax.text(50, 7, "front board/flex split not frozen", fontsize=5.6, ha="center")


def draw_camera_stack_view(ax) -> None:
    setup_view(ax, "5 Camera/lens Z-stack", (0, 42), (0, 24), grid=True)
    ax.set_xlabel("module width / mm", fontsize=6)
    ax.set_ylabel("Z / mm", fontsize=6)
    view_box(ax, 4.5, 3, 32.5, 3, "sensor PCB candidate", MECH_COLORS["camera"], fontsize=5.5)
    view_box(ax, 17, 6, 8.5, 12, "lens\nHOLD", "#ced4da", fontsize=5.5)
    view_box(ax, 5, 1, 25, 1.2, "FPC tail", MECH_COLORS["fpc"], fontsize=5.2)
    view_dim(ax, 39, 0, 39, 18, "18 mm\nRadxa baseline\nnot schematic-frozen", fontsize=5.2)


def draw_temple_top_view(ax, side: str) -> None:
    is_right = side == "right"
    setup_view(ax, "10 Right temple top view" if is_right else "6 Left temple top view", (-3, 153 if is_right else 148), (-3, 24 if is_right else 20), grid=True)
    ax.set_xlabel("length from hinge / mm", fontsize=6)
    ax.set_ylabel("internal height / mm", fontsize=6)
    if is_right:
        view_box(ax, 0, 0, 150, 20, "", MECH_COLORS["frame"], lw=1.3)
        view_box(ax, 0, 0, 8, 20, "FPC\n8", MECH_COLORS["fpc"], fontsize=5.2)
        view_box(ax, 8, 1, 62, 18, "Compute PCB\n62 x 18", MECH_COLORS["compute"], fontsize=6.3)
        view_box(ax, 11, 1.4, 22, 17.2, "U1\nRK3576", "#48cae4", fontsize=5.6)
        view_box(ax, 34, 12, 18, 6.5, "U2 RK806\n+ passives", "#74c0fc", fontsize=5)
        view_box(ax, 34, 2, 12, 10, "U3\nLPDDR", "#90dbf4", fontsize=5.2)
        view_box(ax, 47, 2, 12, 9, "U4\neMMC", "#90dbf4", fontsize=5.2)
        view_box(ax, 52.5, 12, 10.5, 6.5, "U6/L1\nboost", "#74c0fc", fontsize=5)
        view_box(ax, 55.5, 3.9, 13.0, 12.2, "U11\nFCU760K", "#bde0fe", fontsize=5.4)
        view_box(ax, 74, 4, 70, 12.8, "BT1 LP451165\n70 x 12.8 x 5.6", MECH_COLORS["battery"], fontsize=6.2)
        view_box(ax, 108, 0.4, 16, 3, "LS1 slot", MECH_COLORS["speaker"], fontsize=5.5)
        view_box(ax, 136, 0.8, 9, 3.2, "pogo", MECH_COLORS["pogo"], fontsize=5.5)
        view_box(ax, 126, 17.1, 22, 2.4, "Wi-Fi ant.\nkeep-out", MECH_COLORS["rf"], hatch="//", fontsize=4.8)
        view_dim(ax, 0, 22, 150, 22, "usable >=150 mm")
    else:
        view_box(ax, 0, 0, 145, 16, "", MECH_COLORS["frame"], lw=1.3)
        view_box(ax, 0, 0, 8, 16, "FPC\n8", MECH_COLORS["fpc"], fontsize=5.2)
        view_box(ax, 8, 1, 46, 14, "AON/Power PCB\n46 x 14", MECH_COLORS["aon"], fontsize=6.3)
        view_box(ax, 11, 8, 8, 6, "nRF54", "#95d5b2", fontsize=5)
        view_box(ax, 21, 8, 8, 6, "nPM1300", "#95d5b2", fontsize=5)
        view_box(ax, 31, 8, 8, 6, "NDP120\nHOLD", "#95d5b2", fontsize=4.8)
        view_box(ax, 41, 8, 4.5, 4, "BMI270", "#95d5b2", fontsize=4.6)
        view_box(ax, 12, 2, 18, 4.5, "shunts/prot.\nINA pads", "#d8f3dc", fontsize=5)
        view_box(ax, 60, 1, 70, 12.8, "BT2 LP451165\n70 x 12.8 x 5.6", MECH_COLORS["battery"], fontsize=6.2)
        view_box(ax, 122, 14.1, 21, 1.7, "BLE ant.\nkeep-out", MECH_COLORS["rf"], hatch="//", fontsize=4.8)
        view_box(ax, 134, 0.6, 8.5, 3, "pogo", MECH_COLORS["pogo"], fontsize=5.5)
        view_dim(ax, 0, 18, 145, 18, "usable >=145 mm")
    ax.plot([6, 60 if not is_right else 74, 144 if is_right else 130], [0.35, 0.35, 0.35], color="#495057", linewidth=1.7)


def draw_temple_side_view(ax, side: str) -> None:
    is_right = side == "right"
    title = "11 Right temple side view" if is_right else "7 Left temple side view"
    setup_view(ax, title, (-3, 153 if is_right else 148), (-0.8, 10.2), grid=True)
    ax.set_xlabel("length from hinge / mm", fontsize=6)
    ax.set_ylabel("thickness / mm", fontsize=6)
    length = 150 if is_right else 145
    height = 8.5 if is_right else 7.5
    pcb_len = 62 if is_right else 46
    pcb_color = MECH_COLORS["compute"] if is_right else MECH_COLORS["aon"]
    view_box(ax, 0, 0, length, height, "", MECH_COLORS["frame"], lw=1.3)
    view_box(ax, 8, height - 3.2, pcb_len, 1.0, "PCB", pcb_color, fontsize=5.6)
    if is_right:
        view_box(ax, 11, height - 2.2, 22, 1.6, "RK3576", "#48cae4", fontsize=5.2)
        view_box(ax, 34, height - 2.2, 29, 1.6, "DDR/PMIC/boost", "#74c0fc", fontsize=5.2)
        view_box(ax, 74, 1, 70, 5.6, "BT1 max + foam/swell", MECH_COLORS["battery"], fontsize=5.8)
        view_box(ax, 108, height - 1.9, 16, 1.6, "speaker", MECH_COLORS["speaker"], fontsize=5.2)
        ax.plot([8, 70], [height - 0.25, height - 0.25], color=MECH_COLORS["heat"], linewidth=2)
        ax.text(10, height + 0.15, "battery is longitudinally separated from compute heat", fontsize=5.4, color=MECH_COLORS["heat"])
    else:
        view_box(ax, 11, height - 2.2, 34, 1.6, "AON ICs + charger/DSP", "#95d5b2", fontsize=5.2)
        view_box(ax, 60, 1, 70, 5.6, "BT2 max + foam/swell", MECH_COLORS["battery"], fontsize=5.8)
        ax.text(10, height + 0.15, "left side view now explicit; max heights still supplier-HOLD", fontsize=5.4, color="#495057")
    view_dim(ax, length + 1, 0, length + 1, height, f">={height} mm")


def draw_cross_section(ax, title: str, *, pcb_color: str, component_label: str | None = None, battery: bool = False, right_heat: bool = False) -> None:
    setup_view(ax, title, (0, 28), (0, 14), grid=True)
    ax.set_xlabel("width / mm", fontsize=5.5)
    ax.set_ylabel("Z / mm", fontsize=5.5)
    view_box(ax, 2, 1, 24, 12, "", MECH_COLORS["shell"], lw=1.1)
    ax.plot([4, 24], [11.8, 11.8], color="#333", linewidth=1.2)
    ax.text(14, 12.2, "outer wall", fontsize=4.8, ha="center")
    if right_heat:
        ax.plot([6, 22], [10.8, 10.8], color=MECH_COLORS["heat"], linewidth=2)
        ax.text(14, 11.1, "heat spreader", fontsize=4.8, ha="center", color=MECH_COLORS["heat"])
    view_box(ax, 5, 7.2, 18, 0.8, "PCB", pcb_color, fontsize=4.8)
    if component_label:
        view_box(ax, 9, 8.0, 8, 2.4, component_label, "#74c0fc" if pcb_color == MECH_COLORS["compute"] else "#95d5b2", fontsize=4.8)
    if battery:
        view_box(ax, 6, 2.2, 16, 4.2, "battery\n5.6 max env.", MECH_COLORS["battery"], fontsize=5)
        ax.text(14, 1.2, "foam / insulation / swell allowance", fontsize=4.7, ha="center")


def draw_hinge_view(ax, title: str, *, closed: bool) -> None:
    setup_view(ax, title, (0, 70), (0, 34), grid=False)
    view_box(ax, 5, 14, 23, 7, "left\nAON", MECH_COLORS["aon"], fontsize=5.5)
    view_box(ax, 42, 14, 23, 7, "right\ncompute", MECH_COLORS["compute"], fontsize=5.5)
    if closed:
        ax.plot([28, 34, 42], [17, 10, 17], color="#495057", linewidth=3)
        ax.text(35, 6, "closed: no pinch, no vias/pads in bend", fontsize=5.2, ha="center")
    else:
        ax.plot([28, 34, 42], [17, 25, 17], color="#495057", linewidth=3)
        ax.text(35, 6, "open: bend radius, current rating, life test TBD", fontsize=5.2, ha="center")
    ax.text(35, 29, "J4 custom FPC 6-10 mm wide; HOLD", fontsize=5.4, ha="center")


def draw_mechanical_views(path_base: Path) -> tuple[Path, Path, Path]:
    FIGS.mkdir(parents=True, exist_ok=True)
    fig = plt.figure(figsize=(17.0, 24.0), dpi=190)
    gs = fig.add_gridspec(6, 4, height_ratios=[1.15, 1.18, 0.92, 0.82, 0.82, 0.82], hspace=0.58, wspace=0.32)

    draw_whole_top_view(fig.add_subplot(gs[0, 0:2]))
    draw_whole_side_view(fig.add_subplot(gs[0, 2:4]))
    draw_temple_top_view(fig.add_subplot(gs[1, 0:2]), "right")
    draw_temple_top_view(fig.add_subplot(gs[1, 2:4]), "left")
    draw_temple_side_view(fig.add_subplot(gs[2, 0:2]), "right")
    draw_temple_side_view(fig.add_subplot(gs[2, 2:4]), "left")
    draw_front_front_view(fig.add_subplot(gs[3, 0]))
    draw_front_top_view(fig.add_subplot(gs[3, 1]))
    draw_camera_stack_view(fig.add_subplot(gs[3, 2]))
    draw_hinge_view(fig.add_subplot(gs[3, 3]), "15 Hinge/FPC open-state section", closed=False)
    draw_cross_section(fig.add_subplot(gs[4, 0]), "12 Right A-A at RK3576", pcb_color=MECH_COLORS["compute"], component_label="RK3576", right_heat=True)
    draw_cross_section(fig.add_subplot(gs[4, 1]), "13 Right B-B Boost/PMIC", pcb_color=MECH_COLORS["compute"], component_label="boost/\nPMIC", right_heat=True)
    draw_cross_section(fig.add_subplot(gs[4, 2]), "14 Right C-C battery", pcb_color=MECH_COLORS["compute"], battery=True)
    draw_hinge_view(fig.add_subplot(gs[4, 3]), "16 Hinge/FPC closed-state section", closed=True)
    draw_cross_section(fig.add_subplot(gs[5, 0]), "8 Left A-A at tallest AON component", pcb_color=MECH_COLORS["aon"], component_label="AON\nIC")
    draw_cross_section(fig.add_subplot(gs[5, 1]), "9 Left B-B at battery", pcb_color=MECH_COLORS["aon"], battery=True)
    ax_note = fig.add_subplot(gs[5, 2:4])
    setup_view(ax_note, "Legend / gate status", (0, 100), (0, 40), grid=False)
    legend = [
        (MECH_COLORS["compute"], "compute PCB / RK3576 island"),
        (MECH_COLORS["aon"], "AON/power PCB"),
        (MECH_COLORS["battery"], "LP451165 max envelope 70 x 12.8 x 5.6"),
        (MECH_COLORS["rf"], "antenna keep-out, still G14 HOLD"),
        (MECH_COLORS["heat"], "outer-wall heat spreader path"),
    ]
    for i, (color, text) in enumerate(legend):
        y = 31 - i * 5.2
        view_box(ax_note, 5, y, 8, 3, "", color, hatch="//" if "antenna" in text else None)
        ax_note.text(17, y + 1.5, text, va="center", fontsize=7)
    ax_note.text(5, 4, "Placement envelopes only. Formal routing/fab release remains HOLD until P0 supplier data, CAD, RF, battery and J4 close.", fontsize=7)

    fig.suptitle("AI Glasses V2.3 Mechanical Views - 16-view placement envelope atlas", fontsize=16, fontweight="bold", x=0.02, ha="left")
    fig.text(0.02, 0.012, "Style updated from Phase 1.5 floorplan: mm axes, larger temple views, explicit side sections. Source: hardware_mechanical_database.yaml + current schematic/BOM. Not released PCB geometry.", fontsize=8)
    png = path_base.with_suffix(".png")
    svg = path_base.with_suffix(".svg")
    pdf = path_base.with_suffix(".pdf")
    fig.savefig(png, bbox_inches="tight")
    fig.savefig(svg, bbox_inches="tight")
    fig.savefig(pdf, bbox_inches="tight")
    plt.close(fig)
    return svg, png, pdf


def write_pdf(records: list[dict[str, Any]], mech_pdf: Path) -> Path:
    pdf = OUT / "AI_Glasses_Hardware_Dimensions_and_Performance_V2.3.pdf"
    OUT.mkdir(parents=True, exist_ok=True)
    major = ordered_major_records(records)
    with PdfPages(pdf) as pp:
        # Cover/summary page.
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        ax.axis("off")
        ax.text(0.03, 0.92, "AI Glasses V2.3 Hardware Dimensions and Performance", fontsize=20, fontweight="bold", transform=ax.transAxes)
        ax.text(0.03, 0.86, f"Generated {TODAY} from current schematic/BOM and hardware_mechanical_database.yaml", fontsize=10, transform=ax.transAxes)
        summary = [
            "Overall verdict before PCB layout: HOLD",
            "Thin optical frame: FAIL for current architecture",
            "Thick smart-glasses / sports frame: PASS WITH CONDITIONS after CAD/FPC/RF/battery proof",
            "AON and compute PCBs contain multiple ICs because those ICs are soldered to the board; this is allowed",
            "Real overlap of 3D envelopes, courtyards, battery swell, speaker cavity, or RF keep-out is not allowed",
            "Existing ERC report: 0 errors / 0 warnings; DRC not run because V2 has no PCB yet",
        ]
        y = 0.76
        for s in summary:
            ax.text(0.06, y, "- " + s, fontsize=11, transform=ax.transAxes)
            y -= 0.06
        ax.text(0.03, 0.28, "Current power model", fontsize=14, fontweight="bold", transform=ax.transAxes)
        power_rows = [
            ("Deep Off / AON", "22 mW", "20-50 mW target"),
            ("Quick Standby", "134 mW", "DDR/light sleep model"),
            ("Phone assist", "362 mW", "3-5 h blended target"),
            ("Mixed motion", "450 mW", "<=500-800 mW target"),
            ("1080p record", "1290 mW", "45-90 min product target"),
            ("AI + record", "2065 mW", "30-60 min / short burst target"),
        ]
        table = ax.table(cellText=power_rows, colLabels=["State", "Model", "Target/use"], loc="lower left", bbox=[0.03, 0.03, 0.65, 0.20])
        table.auto_set_font_size(False)
        table.set_fontsize(8.5)
        pp.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # Mechanical views page.
        fig, ax = plt.subplots(figsize=(11.69, 16.53))
        ax.axis("off")
        img = plt.imread(mech_pdf.with_suffix(".png"))
        ax.imshow(img)
        ax.set_title("16 required mechanical views - placement envelope atlas", fontsize=13, fontweight="bold")
        pp.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # Design review page.
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        ax.axis("off")
        ax.text(0.03, 0.95, "Phase 1 design review answers", fontsize=16, fontweight="bold", transform=ax.transAxes)
        bullets = [
            "Left temple side view is missing in the earlier render: HOLD until left side + A-A/B-B sections exist in CAD.",
            "Right temple side section is directionally correct but incomplete: add real max heights, solder, heat pad, insulation and tolerance stack.",
            "nRF54L15, nPM1300, NDP120 and BMI270 on one AON PCB is reasonable, with RF/thermal/vibration separation.",
            "AON 46 x 14 mm and Compute 62 x 18 mm are placement envelopes, not proof of complete circuitry.",
            "RK3576 + LPDDR routing needs 8-10 layer HDI review, ball maps, stack-up and DDR length report.",
            "Battery must stay longitudinally separated from RK3576/PMIC/boost; any real hot-source overlap is FAIL.",
            "Wi-Fi/BLE antenna keep-outs are HOLD until antenna SKU/tune and speaker/battery/screw clearance are proven.",
            "Cross-temple 1S2P and J4 current path are major HOLD items.",
        ]
        y = 0.86
        for b in bullets:
            ax.text(0.05, y, "- " + textwrap.fill(b, 115), fontsize=9.3, transform=ax.transAxes, va="top")
            y -= 0.085
        pp.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # KiCad audit evidence page.
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        ax.axis("off")
        ax.text(0.03, 0.95, "KiCad audit evidence generated", fontsize=16, fontweight="bold", transform=ax.transAxes)
        evidence_rows = [
            ["Components", str(len(COMPONENTS)), "symbol_to_footprint_mapping.csv maps each schematic component to current footprint/status"],
            ["Schematic nets", str(len(all_nets())), "interconnect_matrix.csv/md records endpoints, domain crossings and layout rules"],
            ["Footprint release", "HOLD", "VERIFY/HOLD/TBD/DNP items are NOT_FOR_FAB until official land patterns and package heights exist"],
            ["ERC", "Existing PASS", "erc.report.txt shows 0 errors / 0 warnings on 2026-07-01T01:26:33; not rerun because kicad-cli unavailable"],
            ["DRC / 3D", "NOT RUN", "V2 currently has no .kicad_pcb or verified STEP assembly"],
        ]
        table = ax.table(cellText=evidence_rows, colLabels=["Audit item", "Result", "Evidence"], loc="upper left", bbox=[0.03, 0.38, 0.92, 0.48])
        table.auto_set_font_size(False)
        table.set_fontsize(8.0)
        ax.text(0.03, 0.28, "Decision", fontsize=13, fontweight="bold", transform=ax.transAxes)
        ax.text(
            0.05,
            0.21,
            textwrap.fill(
                "No KiCad schematic, footprint, PCB or 3D-model files were modified in this phase. "
                "That is intentional: the project rules require HOLD placeholders until exact MPNs, "
                "official package drawings, pinouts, land patterns and mechanical envelopes are available.",
                120,
            ),
            fontsize=10,
            transform=ax.transAxes,
            va="top",
        )
        pp.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # Component cards, 4 per page.
        for i in range(0, len(major), 4):
            fig, axes = plt.subplots(2, 2, figsize=(11.69, 8.27))
            for ax, r in zip(axes.flat, major[i : i + 4], strict=False):
                draw_package(ax, r)
            for ax in axes.flat[len(major[i : i + 4]) :]:
                ax.axis("off")
            fig.suptitle("Component cards - simple language, source-bound dimensions", fontsize=13, fontweight="bold")
            pp.savefig(fig, bbox_inches="tight")
            plt.close(fig)

        # Open items page.
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        ax.axis("off")
        ax.text(0.03, 0.95, "P0 open mechanical data before PCB layout", fontsize=16, fontweight="bold", transform=ax.transAxes)
        p0 = [r for r in OPEN_ITEMS if r[-1] == "P0"][:14]
        y = 0.88
        for item in p0:
            line = f"{item[0]}: {item[2]}; closure: {item[7]}"
            ax.text(0.05, y, "- " + textwrap.fill(line, 115), fontsize=8.5, transform=ax.transAxes, va="top")
            y -= 0.058
        pp.savefig(fig, bbox_inches="tight")
        plt.close(fig)
    return pdf


def ordered_major_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_ref = {r["reference_designator"]: r for r in records}
    ordered: list[dict[str, Any]] = []
    seen: set[str] = set()
    for ref in COMPONENT_CARD_REFS:
        if ref in seen or ref not in by_ref:
            continue
        ordered.append(by_ref[ref])
        seen.add(ref)
    return ordered


def main() -> None:
    records = build_database()
    yml, csv_path = write_database(records)
    design_review = write_design_review(records)
    open_md, freeze_csv, freeze_xlsx = write_open_items()
    footprint = write_footprint_audit(records)
    symbol_mapping = write_symbol_to_footprint_mapping(records)
    interconnect_csv, interconnect_md = write_interconnect_matrix()
    status_paths = write_phase_and_status(records)
    review_paths = write_engineering_reviews()
    perf_md = write_performance_report(records)
    mech_svg, mech_png, mech_pdf = draw_mechanical_views(FIGS / "AI_Glasses_V2_3_Mechanical_Views")
    perf_pdf = write_pdf(records, mech_pdf)
    generated = [
        yml,
        csv_path,
        design_review,
        open_md,
        freeze_csv,
        freeze_xlsx,
        footprint,
        symbol_mapping,
        interconnect_csv,
        interconnect_md,
        *status_paths,
        *review_paths,
        perf_md,
        mech_svg,
        mech_png,
        mech_pdf,
        perf_pdf,
    ]
    print("Generated mechanical freeze pack:")
    for p in generated:
        print(f"- {p.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
